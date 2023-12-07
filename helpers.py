'''
Nova Home Support Data Processor
filename: helpers.py
This file contains all the helper functions for the payroll-processing web application.
'''

# Import Python packages
import numpy as np
import pandas as pd # For importing, manipulating, and exporting data
import re # Python regular expression support
import datetime # Python datetime conversion support
from dateutil.easter import *
import holidays
from copy import deepcopy
from collections import namedtuple
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Font, Side
from openpyxl.styles import colors, borders, numbers
from openpyxl.styles.borders import Border
from pandas.api.types import CategoricalDtype
from openpyxl.utils import get_column_letter
import os


def test():
    '''
    Test helpers.py
    '''
    return "Hello World"

def delete_files_in_folder(folder_path):
    '''
    Delete all files in a folder path (for refreshing the webapp).

    folder_path -- path to folder
    '''
    try:
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            if os.path.isfile(file_path):
                os.remove(file_path)
        return True, None  # Successful deletion
    except Exception as e:
        return False, str(e)  # Error occurred

def get_name_list(shift_record_path, old_tracker_path):
    '''
    Get a list of names of non-managers from shift record

    shift_record_path -- path to folder
    old_tracker_path -- path to old tracker

    '''
    # Don't do anything if there's something wrong with the uploaded files.
    # Problem with uploaded files shoudl be flagged in other functions.
    try:
        # get the names
        df = pd.read_excel(shift_record_path)
        df = df[['Service Provider']]
        df['Name'] = df['Service Provider'].str.split(' /', n=1).str[0]
        df[['Last Name', 'First Name']] = df['Name'].str.split(', ', expand=True)
        # concatenate First name and Last Name columns in the desired order
        df['Name'] = df['First Name'] + ' ' + df['Last Name']
        name_set = set(df['Name'].unique())
        manager_rates = pd.read_excel(old_tracker_path, sheet_name="MANAGER INFO")
        # do not include managers
        manager_set = set(manager_rates['Name'].unique())
        name_set = name_set - manager_set
        name_list = sorted(list(name_set))
        return name_list
    except: 
        return [] 

def check_shift_overlap(df, name):
    '''
    Check overlapping shifts for the same person.

    df -- a pandas dataframe containing shift records
    name -- name of the employee

    algorithm outline:
        1. subsetting to one individual's record
        2. sort by the person's shift by datetime 
        3. check overlap of shift indexed by t and t+1 for all t<T

    returns a string indicating overlapping shifts for the employee.
    '''
    #get all shifts from one person
    df_indiv = df.loc[df.Name==name]
    df_indiv = df_indiv.sort_values(by='CIDT')
    err_string = ""
    if len(df_indiv) == 1: #no overlap if just one shift
        pass
    else:
        for index in range(len(df_indiv)-1):
            if (df_indiv.iloc[index].CODT - df_indiv.iloc[index+1].CIDT).total_seconds() > 60: #considered overlap if overlapping time is greater than 1 minute.
                problem_date = df_indiv.iloc[index+1]['Check-In Date']
                err_string = err_string + (f'Overlapping shifts detected for {name} on {problem_date} for Shift type {df_indiv.iloc[index].Shift} and {df_indiv.iloc[index+1].Shift}.    ')
    return err_string

def check_unusual_overnight(df):
    '''
    Check and flag unusual overnight shifts.

    df --  a pandas dataframe containing shift records

    Rule:
        if any of the shifts in 'OA1', 'OA2', 'IHSS-Asleep', 'OPA' start before 10:45 PM
        or 7:15 AM, then flag it as an error.

    returns a string indicating unsual overnight shifts for the employee.
    '''
    allowed_shifts = ['OA1', 'OA2', 'IHSS-Asleep', 'OPA']
    start_time = datetime.time(7, 15)  # 7:15 AM
    end_time = datetime.time(22, 45)   # 10:45 PM
    # Create a mask for your conditions
    mask = (
        df['Shift'].isin(allowed_shifts) & 
        (df['CIDT'].dt.time > start_time) & 
        (df['CIDT'].dt.time < end_time)
    )
    # Apply the mask to get a subset of susceptible shifts
    subset_df = df[mask]
    err_string = ""
    for _, row in subset_df.iterrows():
        name = row['Name']
        problem_date = row['Check-In Date']
        problem_shift = row['Shift']
        err_string = err_string + (f'Unusual timing for shift {problem_shift} detected for {name} on {problem_date}.    ')
    return err_string

def approved_holiday(years):
    '''
    Get a set of approved holiday hours.

    years -- a iterable containing the years appeared in the shift record

    return a set of approved holiday dates.
    '''
    # set of holidays
    approved_holiday = set()
    holiday_set = set(["Thanksgiving", "Christmas Day"])
    for year in years:
        holiday_ls = {k for k, v in holidays.US(years=year).items() if v in holiday_set}
        approved_holiday = approved_holiday.union(holiday_ls)
        approved_holiday.add(easter(year)) #Easter
        approved_holiday.add(datetime.date(year, 12, 24)) #Christmas eve
        approved_holiday.add(datetime.date(year, 12, 31)) #NewYear
        approved_holiday.add(datetime.date(year-1, 12, 31)) #Last NewYear
    return approved_holiday

def approved_holiday_hours(years):
    '''
    Return a list of approve holiday datetime range.

    years -- a iterable containing the years appeared in the shift record

    return a list of holiday datetime ranges
    '''
    #create a time interval object
    Range = namedtuple('Range', ['start', 'end'])
    approved_holiday_dt = []
    #append holiday hours according to Nova's unique rule.
    for x in approved_holiday(years):
        if (x.month == 12 and x.day == 31):
            approved_holiday_dt.append(Range(start = datetime.datetime.combine(x, datetime.time(hour=15)),end = datetime.datetime.combine(x+datetime.timedelta(days=1), datetime.time(hour=0))))
        else:
            approved_holiday_dt.append(Range(start = datetime.datetime.combine(x, datetime.time(hour=0)), end = datetime.datetime.combine(x+datetime.timedelta(days=1), datetime.time(hour=0))))
    return approved_holiday_dt

def work_holiday_overlap(work_range, ahh_list):
    '''
    Get the overlap of worked hours and approved holiday hours

    work_range -- the range of the worked shifts during the pay cycle
    ahh_list -- a list of approved holiday hours (datetime range)

    return the total overlapping time in minutes.
    '''
    total_overlap = 0.0
    for ahh in ahh_list:
        latest_start = max(work_range.start, ahh.start)
        earliest_end = min(work_range.end, ahh.end)
        delta = (earliest_end - latest_start)
        overlap = max(0, delta.total_seconds())
        total_overlap += overlap
    return total_overlap/60 #in minutes

def calc_worked_holiday(df):
        '''
        Calculate the # holiday hours each shift contains and add the information to the dataframe.

        df -- a pandas dataframe containing shift records

        return a pandas dataframe with holiday worked hours
        '''
        #create a time interval object
        Range = namedtuple('Range', ['start', 'end'])
        # a list of all years in the shift data
        years = set([min([df.iloc[i]['CIDT'].year for i in range(len(df))]), max([df.iloc[i]['CODT'].year for i in range(len(df))])])
        #range for each shift
        duration_range = [Range(start = df.iloc[i]['CIDT'].to_pydatetime(), end = df.iloc[i]['CODT'].to_pydatetime()) for i in range(len(df))]
        ahh_list = approved_holiday_hours(years)
        holiday_work_time = [work_holiday_overlap(duration_range[i], ahh_list) for i in range(len(df))]
        df['Holiday Worked Duration (Minutes)'] = holiday_work_time
        return df

def is_manager(name, manager_rates):
    '''
    Check if the person is a manager

    name -- name of the employee
    manager_rates -- pandas dataframe containing manager rates.

    return a boolean indicating manager status.
    '''
    return name in set(manager_rates['Name'])

def manager_is_exempt(name, df):
    '''
    Check if the manager has exempt status in that week, assuming the dataframe is a cleaned weekly df.
    
    name -- name of the employee
    df -- a pandas dataframe containing shift records

    return a boolean indicating exempt status.
    '''
    exempt_mins = df.loc[(df['Name'] == name) & (df['Shift'] != 'MGR-Direct-Care')]['Min. Worked'].sum()
    non_exempt_mins = df.loc[(df['Name'] == name) & (df['Shift'] == 'MGR-Direct-Care')]['Min. Worked'].sum()
    return exempt_mins >= non_exempt_mins

def worked_overtime(name, df):
    '''
    Check if a person worked overtime in that week, assuming the dataframe is a cleaned weekly df.
    
    name -- name of the employee
    df -- a pandas dataframe containing shift records for one work week

    return a boolean indicating whether the employee worked overtime for the week
    '''
    return df.loc[df['Name'] == name]['Min. Worked'].sum() > (40*60)

def split_by_work_week(df):
    '''
    Split shift records into multiple dataframes by work week.
    This function assumes that cross-week shifts have been split in two
        
    df -- a pandas dataframe containing shift records

    return a list of dataframe in which each element is a dataframe containing shift record for each week.
    '''
    # Convert the CIDT column to a datetime datatype
    df['CIDT'] = pd.to_datetime(df['CIDT'])
    # Set the CIDT column as the dataframe index
    df.set_index('CIDT', inplace=True)
    # Resample the dataframe by work week (Mon-Sun)
    week_groups = df.resample('W-MON', label='left', closed='left')
    # Create a dictionary to store each work week dataframe
    week_dataframes = {}
    # Iterate over the work week groups and store each dataframe in the dictionary
    for week_start, week_df in week_groups: 
        week_df.reset_index(inplace=True) 
        week_dataframes[week_start.strftime('%Y-%m-%d')] = week_df 
    df.reset_index(inplace=True) 
    return week_dataframes 

def read_shift_record(shift_record_path):
    '''
    Read in shift records, check for errors, and return cleaned dataset along with other relevant info.

    shift_record_path -- file path to shift record

    return a pandas dataframe df, and strings PAY_PERIOD, start_date, end_date 
    '''
    #read dataset from path
    try:
        df = pd.read_excel(shift_record_path)
    except:
        raise FileNotFoundError("Cannot open the shift record file. Make sure it's an .XLSX file.")
    try:
        report_criteria = pd.read_excel(shift_record_path, sheet_name="Report Criteria", engine='openpyxl')
        # Convert the dataframe into a dictionary for easier access
        criteria_dict = dict(zip(report_criteria['Report Criteria'], report_criteria['Value']))
        # Parse the dates and create datetime.datetime objects
        start_date = datetime.datetime.strptime(criteria_dict.get('Slot Start Date From'), '%m/%d/%Y')
        end_date = datetime.datetime.strptime(criteria_dict.get('Slot Start Date To'), '%m/%d/%Y')
        PAY_PERIOD = str(start_date.date()) + ' - ' + str(end_date.date())
        end_date = end_date + datetime.timedelta(days=1)
    except:
        raise ValueError(f"Cannot read report criteria.")
    # indicating whether the data has been pre-processed
    pre_cleaned = False
    #subsetting useful columns
    try: #raw dataset
        df = df[['Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time',
                 'Updated Check-In Date','Updated Check-In Time','Check-Out Date','Check-Out Time','Updated Check-Out Date',
                'Updated Check-Out Time','Staff Worked Duration','Staff Worked Duration (Minutes)']]
    except:
        try: # cleaned dataset
            df = df[['Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time','Check-Out Date',
                     'Check-Out Time','Staff Worked Duration (Minutes)']]
        except:
            raise ValueError("The shift record does not contain all columns needed. It needs to at least contain 'Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time','Check-Out Date','Check-Out Time','Staff Worked Duration (Minutes)'.")
        pre_cleaned = True
    #missing value check
    for col in ['Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time','Check-Out Date','Check-Out Time'
                ,'Staff Worked Duration (Minutes)']:
        if len(df.loc[df[col].isnull()]) != 0:
            raise ValueError(f"Some shifts have missing {col}.")
    #clean shift code
    df['Shift'] = df['Service 1 Description (Code)'].str.replace(r'\(.*\)', '')
    # Remove prefix if it exists
    prefix = 'RC-SDP-CLS-320 '
    df['Shift'] = df['Shift'].apply(lambda x: x[len(prefix):] if x.startswith(prefix) else x)
    df['Shift'] = df['Shift'].apply(lambda x: x.rstrip())
    #clean name
    df['Name'] = df['Service Provider'].str.split(' /', n=1).str[0]
    df[['Last Name', 'First Name']] = df['Name'].str.split(', ', expand=True)
    # concatenate First name and Last Name columns in the desired order
    df['Name'] = df['First Name'] + ' ' + df['Last Name']
    # drop First name and Last Name columns
    df = df.drop(["Service 1 Description (Code)", "Service Provider"], axis=1)
    # Replace Date/Time with Updated Date/Time if the latter is not NaN
    try:
        df['Check-In Date'] = df['Updated Check-In Date'].fillna(df['Check-In Date'])
        df['Check-In Time'] = df['Updated Check-In Time'].fillna(df['Check-In Time'])
        df['Check-Out Date'] = df['Updated Check-Out Date'].fillna(df['Check-Out Date'])
        df['Check-Out Time'] = df['Updated Check-Out Time'].fillna(df['Check-Out Time'])
        # drop unnecessary columns once the info is integrated.
        df.drop(['Updated Check-In Date', 'Updated Check-In Time','Updated Check-Out Date', 
                 'Updated Check-Out Time'], axis=1, inplace=True)
    except:
        if pre_cleaned == False:
            raise RuntimeError("Cannot update check-in/out date/times using updated check-in/out dates/times.")
    #Convert date and time to python-style datetime format
    try:
        CIDT = df['Check-In Date'].str.cat(df['Check-In Time'], sep=' ')
        CODT = df['Check-Out Date'].str.cat(df['Check-Out Time'], sep=' ')
        CIDT = CIDT.apply(lambda x: datetime.datetime.strptime(x, r'%m/%d/%Y %I:%M %p'))
        CODT = CODT.apply(lambda x: datetime.datetime.strptime(x, r'%m/%d/%Y %I:%M %p'))
        # add python-format datetime to the dataframe
        df['CIDT'] = CIDT
        df['CODT'] = CODT
    except:
        raise RuntimeError("Cannot convert date and time to Python's own format.")
    # Check for overlapping shifts
    err_string = ""
    for name in df.Name.unique():
        err_string = err_string + check_shift_overlap(df, name)
    if err_string != "":
        raise ValueError(err_string)
    # Check for unsual overnight shifts
    err_string = err_string + check_unusual_overnight(df)
    if err_string != "":
        raise ValueError(err_string)
    # Split shifts that span two days
    new_rows = []
    # loop through each row of the original dataframe
    for index, row in df.iterrows():
        # check if CIDT is on the same date as CODT
        if row['CIDT'].weekday() != row['CODT'].weekday():
            # create a new row for the portion of the shift that occurred on Sunday
            first_day_row = row.copy()
            first_day_row['CODT'] = pd.to_datetime(str(row['CODT'].date()) + ' 00:00:00')
            new_rows.append(first_day_row)
            # create a new row for the portion of the shift that occurred on Monday
            second_day_row = row.copy()
            second_day_row['CIDT'] = pd.to_datetime(str(row['CODT'].date()) + ' 00:00:00')
            new_rows.append(second_day_row)
        else:
            new_rows.append(row)
    # create a new dataframe from the modified rows
    new_df = pd.DataFrame(new_rows)
    # sort the new dataframe by CIDT
    new_df = new_df.sort_values(by=['CIDT'])
    # reset the index of the new dataframe
    new_df = new_df.reset_index(drop=True)
    new_df['Check-In Date'] = new_df['CIDT'].apply(lambda x: x.strftime('%m/%d/%Y'))
    new_df['Check-In Time'] = new_df['CIDT'].apply(lambda x: x.strftime('%I:%M %p'))
    new_df['Check-Out Date'] = new_df['CODT'].apply(lambda x: x.strftime('%m/%d/%Y'))
    new_df['Check-Out Time'] = new_df['CODT'].apply(lambda x: x.strftime('%I:%M %p'))
    new_df['Min. Worked'] = ((new_df['CODT'] - new_df['CIDT']).dt.total_seconds() / 60).round(2)
    df = new_df
    df['Shift_original'] = df['Shift']
    #filters out shifts of type "Adaptive Behavior Treatment" since Nova doesn't pay for those
    #df = df.loc[~df['Shift'].str.contains('Adaptive Behavior Treatment')]
    df['Shift'] = df['Shift'].replace({'Training-HSS': 'HSS1', 'Training-RBT': 'BST1'})
    #calculate holiday hours
    calc_worked_holiday(df)
    return (df, PAY_PERIOD, start_date, end_date)

def read_one_person_record(shift_record_path, selected_name):
    '''
    Read the record of a specifc staff for off-cycle payroll.

    shift_record_path -- file path to the shift record

    selected_name -- full name of the staff ("First Last")

    return pandas dataframes: shift record (df), pay period, start date, and end date.
    '''
    #read dataset from path
    try:
        df = pd.read_excel(shift_record_path)
    except:
        raise FileNotFoundError("Cannot open the shift record file. Make sure it's an .XLSX file.")
    try:
        report_criteria = pd.read_excel(shift_record_path, sheet_name="Report Criteria", engine='openpyxl')
        # Convert the dataframe into a dictionary for easier access
        criteria_dict = dict(zip(report_criteria['Report Criteria'], report_criteria['Value']))
        # Parse the dates and create datetime.datetime objects
        start_date = datetime.datetime.strptime(criteria_dict.get('Slot Start Date From'), '%m/%d/%Y')
        end_date = datetime.datetime.strptime(criteria_dict.get('Slot Start Date To'), '%m/%d/%Y')
        PAY_PERIOD = str(start_date.date()) + ' - ' + str(end_date.date())
        end_date = end_date + datetime.timedelta(days=1)
    except:
        raise ValueError(f"Cannot read report criteria.")
    # indicating whether the data has been pre-processed
    pre_cleaned = False
    #subsetting useful columns
    try: #raw dataset
        df = df[['Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time',
                 'Updated Check-In Date','Updated Check-In Time','Check-Out Date','Check-Out Time','Updated Check-Out Date',
                'Updated Check-Out Time','Staff Worked Duration','Staff Worked Duration (Minutes)']]
    except:
        try: # cleaned dataset
            df = df[['Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time','Check-Out Date',
                     'Check-Out Time','Staff Worked Duration (Minutes)']]
        except:
            raise ValueError("The shift record does not contain all columns needed. It needs to at least contain 'Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time','Check-Out Date','Check-Out Time','Staff Worked Duration (Minutes)'.")
        pre_cleaned = True
    #clean name and keep only relevant name
    df['Name'] = df['Service Provider'].str.split(' /', n=1).str[0]
    df[['Last Name', 'First Name']] = df['Name'].str.split(', ', expand=True)
    # concatenate First name and Last Name columns in the desired order
    df['Name'] = df['First Name'] + ' ' + df['Last Name']
    df = df.loc[df['Name'] == selected_name]
    #missing value check
    for col in ['Service 1 Description (Code)','Service Provider','Check-In Date','Check-In Time','Check-Out Date','Check-Out Time'
                ,'Staff Worked Duration (Minutes)']:
        if len(df.loc[df[col].isnull()]) != 0:
            raise ValueError(f"Some shifts have missing {col}.")
    #clean shift code
    df['Shift'] = df['Service 1 Description (Code)'].str.replace(r'\(.*\)', '')
    # Remove prefix if it exists
    prefix = 'RC-SDP-CLS-320 '
    df['Shift'] = df['Shift'].apply(lambda x: x[len(prefix):] if x.startswith(prefix) else x)
    df['Shift'] = df['Shift'].apply(lambda x: x.rstrip())
    # drop First name and Last Name columns
    df = df.drop(["Service 1 Description (Code)", "Service Provider"], axis=1)
    # Replace Date/Time with Updated Date/Time if the latter is not NaN
    try:
        df['Check-In Date'] = df['Updated Check-In Date'].fillna(df['Check-In Date'])
        df['Check-In Time'] = df['Updated Check-In Time'].fillna(df['Check-In Time'])
        df['Check-Out Date'] = df['Updated Check-Out Date'].fillna(df['Check-Out Date'])
        df['Check-Out Time'] = df['Updated Check-Out Time'].fillna(df['Check-Out Time'])
        # drop unnecessary columns once the info is integrated.
        df.drop(['Updated Check-In Date', 'Updated Check-In Time','Updated Check-Out Date', 
                 'Updated Check-Out Time'], axis=1, inplace=True)
    except:
        if pre_cleaned == False:
            raise RuntimeError("Cannot update check-in/out date/times using updated check-in/out dates/times.")
    #Convert date and time to python-style datetime format
    try:
        CIDT = df['Check-In Date'].str.cat(df['Check-In Time'], sep=' ')
        CODT = df['Check-Out Date'].str.cat(df['Check-Out Time'], sep=' ')
        CIDT = CIDT.apply(lambda x: datetime.datetime.strptime(x, r'%m/%d/%Y %I:%M %p'))
        CODT = CODT.apply(lambda x: datetime.datetime.strptime(x, r'%m/%d/%Y %I:%M %p'))
        PAY_PERIOD = str(CIDT.min().date()) + ' - ' + str(CIDT.max().date())
        start_date, end_date = PAY_PERIOD.split(' - ')
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + datetime.timedelta(days=1)
        # add python-format datetime to the dataframe
        df['CIDT'] = CIDT
        df['CODT'] = CODT
    except:
        raise RuntimeError("Cannot convert date and time to Python's own format.")
    # Check for overlapping shifts
    err_string = ""
    for name in df.Name.unique():
        err_string = err_string + check_shift_overlap(df, name)
    if err_string != "":
        raise ValueError(err_string)
    # Check for unsual overnight shifts
    err_string = err_string + check_unusual_overnight(df)
    if err_string != "":
        raise ValueError(err_string)
    # Split shifts that span two days
    new_rows = []
    # loop through each row of the original dataframe
    for index, row in df.iterrows():
        # check if CIDT is on the same date as CODT
        if row['CIDT'].weekday() != row['CODT'].weekday():
            # create a new row for the portion of the shift that occurred on Sunday
            first_day_row = row.copy()
            first_day_row['CODT'] = pd.to_datetime(str(row['CODT'].date()) + ' 00:00:00')
            new_rows.append(first_day_row)
            # create a new row for the portion of the shift that occurred on Monday
            second_day_row = row.copy()
            second_day_row['CIDT'] = pd.to_datetime(str(row['CODT'].date()) + ' 00:00:00')
            new_rows.append(second_day_row)
        else:
            new_rows.append(row)
    # create a new dataframe from the modified rows
    new_df = pd.DataFrame(new_rows)
    # sort the new dataframe by CIDT
    new_df = new_df.sort_values(by=['CIDT'])
    # reset the index of the new dataframe
    new_df = new_df.reset_index(drop=True)
    new_df['Check-In Date'] = new_df['CIDT'].apply(lambda x: x.strftime('%m/%d/%Y'))
    new_df['Check-In Time'] = new_df['CIDT'].apply(lambda x: x.strftime('%I:%M %p'))
    new_df['Check-Out Date'] = new_df['CODT'].apply(lambda x: x.strftime('%m/%d/%Y'))
    new_df['Check-Out Time'] = new_df['CODT'].apply(lambda x: x.strftime('%I:%M %p'))
    new_df['Min. Worked'] = ((new_df['CODT'] - new_df['CIDT']).dt.total_seconds() / 60).round(2)
    df = new_df
    df['Shift_original'] = df['Shift']
    #filters out shifts of type "Adaptive Behavior Treatment" since Nova doesn't pay for those
    #df = df.loc[~df['Shift'].str.contains('Adaptive Behavior Treatment')]
    df['Shift'] = df['Shift'].replace({'Training-HSS': 'HSS1', 'Training-RBT': 'BST1'})
    #calculate holiday hours
    calc_worked_holiday(df)
    return (df, PAY_PERIOD, start_date, end_date)

def read_old_tracker(old_tracker_path, start_date):
    '''
    Read the old tracker and check for errors

    old_tracker_path -- path to the old tracker
    start_date -- start date of the pay cycle
     
    return pandas dataframes: manager_rates, non_manager_rates, staff_info, accrued_hrs, bonus_df, prepaid_last_time, unpaid_last_time
    '''
    #read tabs in the spreadsheet
    manager_rates = pd.read_excel(old_tracker_path, sheet_name="MANAGER INFO")
    non_manager_rates = pd.read_excel(old_tracker_path, sheet_name="SHIFT INFO")
    staff_info = pd.read_excel(old_tracker_path, sheet_name="STAFF INFO")
    accrued_hrs = pd.read_excel(old_tracker_path, sheet_name="HRS & ACCRUALS")
    accrued_hrs = accrued_hrs.fillna(0)
    #transformed bonus_df, per_person_bonus_list, original_bonus_df
    bonus_df = bonus = original_bonus_df = pd.read_excel(old_tracker_path, sheet_name='NEW PTO & BONUS INFO') 
    prepaid_last_time = pd.read_excel(old_tracker_path, sheet_name='IGNORE! (Prepaid Shifts)')  
    unpaid_last_time = pd.read_excel(old_tracker_path, sheet_name='IGNORE! (Next Period Shifts)')
    #Staff info formatting.
    staff_info['Days Elapsed Since Hire Date'] = staff_info['Hire Date'].apply(lambda x: max(0, (start_date - x).days))
    staff_info['Hire Date'] = staff_info['Hire Date'].apply(lambda x: x.date())
    manager_rates['Days Elapsed Since Hire Date'] = manager_rates['Hire Date'].apply(lambda x: max(0, (start_date - x).days))
    manager_rates['Hire Date'] = manager_rates['Hire Date'].apply(lambda x: x.date())
    # Admin/Sick/Vacay Wage
    cd_non_manager_rates = deepcopy(non_manager_rates[~non_manager_rates['Shift'].str.contains('-Not-Worked')])
    cd_non_manager_rates['Shift'] = cd_non_manager_rates['Shift'].str.replace('-Worked$', '', regex=True)
    stf_info2 = deepcopy(staff_info)
    schedule_columns = [col for col in stf_info2.columns if col.startswith("# ")]
    rows_to_fill = stf_info2[schedule_columns].isnull().all(axis=1)
    # Fill '# HSS' column with 1 for the matching rows
    stf_info2.loc[rows_to_fill, '# HSS'] = 1
    bst_levels = stf_info2['BST Level'].dropna().str.strip().unique()
    oa_levels = stf_info2['OA Level'].dropna().str.strip().unique()
    hss_levels = stf_info2['HSS Level'].dropna().str.strip().unique()
    # Combine all unique levels
    unique_levels = pd.unique(list(bst_levels) + list(oa_levels) + list(hss_levels))
    # Remove empty string from unique levels
    unique_levels = [level for level in unique_levels if level != '']
    # Create new columns based on unique levels
    for level in unique_levels:
        stf_info2[level] = 0
    # Set the value of the respective level column to 1 if the employee has that level
    for index, row in stf_info2.iterrows():
        for level in unique_levels:
            if row['BST Level'] == level:
                stf_info2.loc[index, level] = row['# BST']
            elif row['OA Level'] == level:
                stf_info2.loc[index, level] = row['# OA']
            elif row['HSS Level'] == level:
                stf_info2.loc[index, level] = row['# HSS']
    # Remove the original BST Level, OA Level, and HSS Level columns and create a new dataframe
    stf_info2 = stf_info2.drop(['BST Level', 'OA Level', 'HSS Level', '# HSS', '# BST', '# OA', 'Hire Date', 
                                'Accrual Rate', 'Days Elapsed Since Hire Date', 'Admin/Sick/Vacay Wage'], axis=1)
    stf_info2 = stf_info2.rename(columns=lambda x: x.replace("# ", ""))
    stf_info2.columns = stf_info2.columns.str.strip()
    stf_info2 = stf_info2.fillna(0)
    stf_info2 = stf_info2.replace(r'^\s*$', 0, regex=True)
    avs_wage = [] #average wage 
    #Calculate regular rate.
    for index, row in stf_info2.iterrows():
        sum_total = 0
        total_hrs = 0
        for shift_name, value in row[1:].items():
            sum_total += value * cd_non_manager_rates.loc[cd_non_manager_rates.Shift == shift_name]['BOT Hourly Wage'].iloc[0]
            total_hrs += value
        avs_wage.append(sum_total/total_hrs)
    staff_info['Admin/Sick/Vacay Wage'] = avs_wage
    #Clean Bonus record 
    bonus_df = bonus_df.drop(['First Name', 'Last Name'], axis=1)
    bonus_df = bonus_df.rename(columns={'Full Name': 'Name'})
    bonus = pd.DataFrame(columns=["Name", "Date", "Bonus Amount"])
    for index, row in bonus_df.iterrows():
        for bon_num in ['Bonus 1', 'Bonus 2', 'Bonus 3', 'Bonus 4']:
            if np.isnan(row[bon_num]) == False:
                name = row['Name']
                date = row[bon_num+' Date']
                amount = row[bon_num]
                bonus = bonus.append({"Name": name, "Date": date, "Bonus Amount": amount},ignore_index=True)
    # Format bonus dataframe
    bonus_df['Premium Pay 1 Check-In Time'] = pd.to_datetime(bonus_df['Premium Pay 1 Check-In Time'], format='%H:%M:%S').dt.time
    bonus_df['Premium Pay 2 Check-In Time'] = pd.to_datetime(bonus_df['Premium Pay 2 Check-In Time'], format='%H:%M:%S').dt.time
    bonus_df['Premium Pay 3 Check-In Time'] = pd.to_datetime(bonus_df['Premium Pay 3 Check-In Time'], format='%H:%M:%S').dt.time
    bonus_df['Premium Pay 4 Check-In Time'] = pd.to_datetime(bonus_df['Premium Pay 4 Check-In Time'], format='%H:%M:%S').dt.time
    # Combine Check-In Date and Check-In Time into Check-In Datetime
    bonus_df['Premium Pay 1 Check-In Datetime'] = pd.to_datetime(bonus_df['Premium Pay 1 Check-In Date']) + pd.to_timedelta(bonus_df['Premium Pay 1 Check-In Time'].astype(str))
    bonus_df['Premium Pay 2 Check-In Datetime'] = pd.to_datetime(bonus_df['Premium Pay 2 Check-In Date']) + pd.to_timedelta(bonus_df['Premium Pay 2 Check-In Time'].astype(str))
    bonus_df['Premium Pay 3 Check-In Datetime'] = pd.to_datetime(bonus_df['Premium Pay 3 Check-In Date']) + pd.to_timedelta(bonus_df['Premium Pay 3 Check-In Time'].astype(str))
    bonus_df['Premium Pay 4 Check-In Datetime'] = pd.to_datetime(bonus_df['Premium Pay 4 Check-In Date']) + pd.to_timedelta(bonus_df['Premium Pay 4 Check-In Time'].astype(str))
    bonus_df['Premium Pay 1 Check-Out Time'] = pd.to_datetime(bonus_df['Premium Pay 1 Check-Out Time'], format='%H:%M:%S').dt.time
    bonus_df['Premium Pay 2 Check-Out Time'] = pd.to_datetime(bonus_df['Premium Pay 2 Check-Out Time'], format='%H:%M:%S').dt.time
    bonus_df['Premium Pay 3 Check-Out Time'] = pd.to_datetime(bonus_df['Premium Pay 3 Check-Out Time'], format='%H:%M:%S').dt.time
    bonus_df['Premium Pay 4 Check-Out Time'] = pd.to_datetime(bonus_df['Premium Pay 4 Check-Out Time'], format='%H:%M:%S').dt.time
    # Combine Check-Out Date and Check-Out Time into Check-Out Datetime
    bonus_df['Premium Pay 1 Check-Out Datetime'] = pd.to_datetime(bonus_df['Premium Pay 1 Check-Out Date']) + pd.to_timedelta(bonus_df['Premium Pay 1 Check-Out Time'].astype(str))
    bonus_df['Premium Pay 2 Check-Out Datetime'] = pd.to_datetime(bonus_df['Premium Pay 2 Check-Out Date']) + pd.to_timedelta(bonus_df['Premium Pay 2 Check-Out Time'].astype(str))
    bonus_df['Premium Pay 3 Check-Out Datetime'] = pd.to_datetime(bonus_df['Premium Pay 3 Check-Out Date']) + pd.to_timedelta(bonus_df['Premium Pay 3 Check-Out Time'].astype(str))
    bonus_df['Premium Pay 4 Check-Out Datetime'] = pd.to_datetime(bonus_df['Premium Pay 4 Check-Out Date']) + pd.to_timedelta(bonus_df['Premium Pay 4 Check-Out Time'].astype(str))
    # Get duration
    bonus_df['Premium Pay 1 Duration'] = pd.to_datetime(bonus_df['Premium Pay 1 Check-Out Datetime']) - pd.to_datetime(bonus_df['Premium Pay 1 Check-In Datetime'])
    bonus_df['Premium Pay 2 Duration'] = pd.to_datetime(bonus_df['Premium Pay 2 Check-Out Datetime']) - pd.to_datetime(bonus_df['Premium Pay 2 Check-In Datetime'])
    bonus_df['Premium Pay 3 Duration'] = pd.to_datetime(bonus_df['Premium Pay 3 Check-Out Datetime']) - pd.to_datetime(bonus_df['Premium Pay 3 Check-In Datetime'])
    bonus_df['Premium Pay 4 Duration'] = pd.to_datetime(bonus_df['Premium Pay 4 Check-Out Datetime']) - pd.to_datetime(bonus_df['Premium Pay 4 Check-In Datetime'])
    duration_columns = ['Premium Pay 1 Duration', 'Premium Pay 2 Duration', 'Premium Pay 3 Duration', 'Premium Pay 4 Duration']
    bonus_df[duration_columns] = bonus_df[duration_columns].fillna(pd.Timedelta(0))
    bonus_df['Premium Pay Hours'] = (
        bonus_df['Premium Pay 1 Duration'] + bonus_df['Premium Pay 2 Duration'] + bonus_df['Premium Pay 3 Duration'] + bonus_df['Premium Pay 4 Duration']).dt.total_seconds()/3600
    return (manager_rates, non_manager_rates, accrued_hrs, bonus_df, bonus, original_bonus_df, staff_info, prepaid_last_time, unpaid_last_time)

def merge_shifts(df, staff_info, manager_rates, non_manager_rates):
    '''
    Merge shift records with information from the tracker, include RBT conversion, Admin wage inclusion, and merging shift with hourly rates.
    
    df -- a pandas dataframe containing shift records
    staff_info -- a pandas dataframe version of the STAFF INFO TAB of the old tracker
    manager_rates -- a pandas dataframe version of the MANAGER INFO TAB of the old tracker
    non_manager_rates -- a pandas dataframe version of the SHIFT INFO TAB of the old tracker

    return merged datafarme df_shift_merged
    '''
    #Convert RBT to BST
    for _, row in df.iterrows():
        if row['Shift'] == 'RBT':
            name = row['Name']
            matching_row = staff_info[staff_info['Name'] == name]
            if len(matching_row) != 1:
                raise ValueError(f"Error: Multiple or no matching rows found for Name '{name}' in staff info.")
            else:
                if pd.isna(matching_row['BST Level'].values[0]) == False:
                    df.loc[_, 'Shift'] = matching_row['BST Level'].values[0]
                else:
                    raise ValueError(f"Error: RBT is not equal to 1 for Name '{name}' in other_rates.")
    df_shift_merged = pd.merge(df, non_manager_rates, how='left', on='Shift')
    for name, accrued in zip(manager_rates['Name'], manager_rates['Accrual Rate']):
        df_shift_merged.loc[(df_shift_merged['Name'] == name), ['Accrual Rate']] = accrued
    #include Admin Wage
    admin_shifts = df_shift_merged[df_shift_merged['Shift'] == 'Admin']
    admin_names = admin_shifts['Name'].tolist()
    # Filter the rows in other_rates where the NAME column matches the names of the Service Providers in the Admin shifts
    admin_rates = staff_info[staff_info['Name'].isin(admin_names)]
    # Merge the Admin rates back into the Admin shifts dataframe
    if len(admin_names)>0:
        admin_shifts_merged = pd.merge(admin_shifts, admin_rates, on='Name', how='left')
        # Fill the Regular Hourly Wage and Overtime Hourly Wage columns with the values from the ADMIN/VACAY WAGE column
        admin_shifts_merged.loc[:, 'Regular Hourly Wage'] = admin_shifts_merged['Admin/Sick/Vacay Wage']
        admin_shifts_merged.loc[:, 'BOT Hourly Wage'] = admin_shifts_merged['Admin/Sick/Vacay Wage']
        admin_shifts_merged.loc[:, 'Accrual Rate'] = 0.04
        df_shift_merged = df_shift_merged[df_shift_merged['Shift'] != 'Admin']
        df_shift_merged = pd.concat([df_shift_merged, admin_shifts_merged], ignore_index=True)
        columns_to_remove = [col for col in df_shift_merged.columns if col.startswith("# ")] + ['Billing Rates', 'Accrual Rate_x', 
                                                                                                'Hire Date', 'BST Level', 'HSS Level', 
                                                                                                'OA Level', 'Accrual Rate_y',
                                                                                                'Days Elapsed Since Hire Date', 'Admin/Sick/Vacay Wage', 
                                                                                                'Staff Worked Duration (Minutes)']
        df_shift_merged = df_shift_merged.drop(columns=columns_to_remove)
    return df_shift_merged

def calc_time_off(df_shift_merged):
    '''
    Collect and Calculate Vacation and Sick time for each employee.

    df_shift_merged -- a pandas dataframe containing shift records merged with tracker information

    return dataframes df_shift_merged, time_off, time_off_as_shifts
    '''
    #time off
    all_names = df_shift_merged['Name'].drop_duplicates()
    #sick leave
    sick_shift_sum = df_shift_merged[df_shift_merged['Shift'] == 'Sick'].groupby('Name')['Min. Worked'].sum().reset_index()
    sick_shift_sum = all_names.to_frame().merge(sick_shift_sum, on='Name', how='left').fillna(0)
    sick_shift_sum = sick_shift_sum.rename(columns={'Min. Worked': 'Sick Hrs'})
    sick_shift_sum['Sick Hrs'] = sick_shift_sum['Sick Hrs']/60
    #vacation cash-out
    vac_shift_sum = df_shift_merged[df_shift_merged['Shift'] == 'Vacation'].groupby('Name')['Min. Worked'].sum().reset_index()
    vac_shift_sum = all_names.to_frame().merge(vac_shift_sum, on='Name', how='left').fillna(0)
    vac_shift_sum = vac_shift_sum.rename(columns={'Min. Worked': 'Vac Hrs'})
    vac_shift_sum['Vac Hrs'] = vac_shift_sum['Vac Hrs']/60
    time_off = vac_shift_sum.merge(sick_shift_sum, on='Name')
    #time_off_as_shifts is a subset of df_shift_merged with only Sick and Vacation in there.
    time_off_as_shifts = df_shift_merged[((df_shift_merged['Shift'] == 'Sick') | (df_shift_merged['Shift'] == 'Vacation'))]
    df_shift_merged = df_shift_merged[~((df_shift_merged['Shift'] == 'Sick') | (df_shift_merged['Shift'] == 'Vacation'))]
    return (df_shift_merged, time_off, time_off_as_shifts)

def crop_shifts(df, start_date, end_date):
    '''
    Crop shift record and keep those that are within this pay cycle.

    df -- a pandas dataframe containing shift records
    start_date, end_date -- start and end dates of the pay period

    return dataframes df, df_after_pay_period, prepaid_hours,week_order and string PREPAY
    '''
    # get shifts that wil be paid in the next pay period
    #To be paid Next period:
    df_after_pay_period = deepcopy(df[df['CIDT'] >= end_date])
    # Filter the DataFrame based on the time duration
    df = df[(df['CIDT'] >= start_date) & (df['CIDT'] < end_date)]
    if len(df) == 0:
        return(df, df_after_pay_period, pd.DataFrame(columns=df.columns),[str(start_date.date())], False)
    ### get shifts that wil be prepaid in this pay cycle (belong to this pay cycle, but falls under the last partial week)
    week_dataframes = split_by_work_week(df)
    week_order = []
    full_week = []
    for week_start, week_df in week_dataframes.items():
        week_order.append(week_start)
        # Get the first and last day of the week
        week_start_day = week_df['CIDT'].min().weekday()
        week_end_day = week_df['CIDT'].max().weekday()
        # Check if the week contains observations for every day from Mon to Sun
        if week_end_day - week_start_day == 6:
            full_week.append(True)
        else:
            full_week.append(False)
    week_df = pd.DataFrame({"week": week_order, "full week": full_week})
    # Convert the "week" column to datetime objects
    week_df['week_dt'] = pd.to_datetime(week_df['week'])
    # Sort the rows by the values in the "week" column
    week_df = week_df.sort_values('week_dt')
    # Drop the first week from week_df if it's not a full week
    if (not week_df.iloc[0]['full week']) and (len(week_dataframes) > 1) :
        # Get the key corresponding to the first row's "week" value
        key_to_drop = week_df.iloc[0]['week']
        # Drop the item with the corresponding key from week_dataframes
        week_dataframes.pop(key_to_drop)
    PREPAY = (week_df.iloc[-1]['full week']==False) # partial week or full week
    prepaid_hours = pd.DataFrame(columns=df.columns)
    if PREPAY:
        prepaid_hours = deepcopy(week_dataframes[week_df.iloc[-1]['week']])
    return (df, df_after_pay_period, prepaid_hours, week_order, PREPAY)

def non_manager_payroll(non_mgr, df_shift_merged, accrued_hrs, bonus_df, bonus, time_off, staff_info, week_order, prepaid_last_time, PAY_PERIOD, new_accrued_hrs):   
    '''
    Process payroll for non-managers and return the result with accrued hours

    non_mgr -- list of non-managers
    df_shift_merged -- shift dataframe for the pay period
    accrued_hrs -- accrued hours (existing for manipulation)
    bonus_df -- bonus info as dataframe
    bonus -- bonus summarized by person
    staff_info -- pay rates for staff
    time_off -- time taken off
    week_order -- order of weeks within the pay period
    prepaid_last_time -- list of shifts prepaid last time
    PAY_PERIOD -- a string of pay period
    new_accrued_hrs -- new accrual dataframe
    '''
    #Create a list that stores the payroll dictionary
    non_mgr_payroll = []
    #For each non-manager
    df_shift_merged = deepcopy(df_shift_merged)
    # name update
    df_shift_merged['Shift'] = df_shift_merged['Shift'].replace({'Training-HSS': 'HSS1', 'Training-RBT': 'BST1'})
    if len(prepaid_last_time) > 0: #who are prepaid last time?
        prepaid_ppl = prepaid_last_time['Name'].unique()
    else:
        prepaid_ppl = set()
    for _, name in enumerate(non_mgr):
        #subset to the individual's shift
        df_indiv = df_shift_merged.loc[df_shift_merged['Name'] == name]
        regular_rate = staff_info.loc[staff_info['Name'] == name]['Admin/Sick/Vacay Wage'].iloc[0]
        #set aggregation rule
        aggregations = {'Min. Worked': 'sum', 'Regular Hourly Wage': 'first', 'Name': 'first'}
        #Aggregate payroll data summarizing total time worked for each shift
        df_payroll = df_indiv[['Shift', 'Min. Worked', 'Regular Hourly Wage', 'Name']]
        df_payroll = df_payroll.groupby('Shift').agg(aggregations)
        df_payroll = df_payroll.reset_index()
        #Calculate total hours worked
        df_indiv_worked = df_indiv[~df_indiv['Shift'].str.contains('-Not-Worked')]
        total_hours_worked = round(df_indiv_worked['Min. Worked'].sum()/60, 2)
        #Dealing with holiday
        if df_indiv_worked['Holiday Worked Duration (Minutes)'].sum() > 0:
            df_holiday_pay = df_indiv_worked[['Name', 'Shift', 'Holiday Worked Duration (Minutes)', 'BOT Hourly Wage']] # filter
            df_holiday_pay = df_holiday_pay[df_holiday_pay['Holiday Worked Duration (Minutes)'] != 0] # with holiday overlap
            df_holiday_pay['Shift'] = df_holiday_pay['Shift'].apply(lambda x: x + ' Holiday Extra Pay') # rename cols
            df_holiday_pay['Regular Hourly Wage'] = (df_holiday_pay['BOT Hourly Wage']*0.5).round(2) #halve the wage for extra pay
            df_holiday_pay = df_holiday_pay.rename(columns={'Holiday Worked Duration (Minutes)': 'Min. Worked'})
            df_holiday_pay = df_holiday_pay.drop('BOT Hourly Wage', axis=1)
            df_holiday_pay = df_holiday_pay.groupby('Shift').agg(aggregations)
            df_holiday_pay = df_holiday_pay.reset_index()
            df_payroll= pd.concat([df_payroll, df_holiday_pay], ignore_index=True)
        #Dealing with weekly overtime pay
        df_weeks = split_by_work_week(df_indiv)
        for key in df_weeks.keys(): #each key is a timestamp
            df_weekly = deepcopy(df_weeks[key])
            if (key == week_order[0]) and (name in prepaid_ppl): #first week and prepaid
                df_weekly = pd.concat([prepaid_last_time.loc[prepaid_last_time.Name == name], df_weekly], ignore_index=True)
            df_weekly_worked = df_weekly[~df_weekly['Shift'].str.contains('-Not-Worked')]
            weekly_hours_worked = round(df_weekly_worked['Min. Worked'].sum()/60, 2)
            weekly_hours_paid = round(df_weekly['Min. Worked'].sum()/60, 2)
            overtime_hours = max(0, weekly_hours_worked-40)
            ot_rate =((df_weekly['BOT Hourly Wage']*(df_weekly['Min. Worked']/60).round(2)).sum()/weekly_hours_paid)
            if overtime_hours > 0:
                #Calculate BOT rate
                df_overtime = pd.DataFrame({'Name': name, 'Shift': [f'OT Extra Pay ({key})'], 'Min. Worked': [overtime_hours*60], 
                                            'Regular Hourly Wage': [round(ot_rate/2, 2)]})
                #Add to payroll
                df_payroll= pd.concat([df_payroll, df_overtime], ignore_index=True)
        # Format df_payroll
        df_payroll = df_payroll.rename(columns={'Regular Hourly Wage': 'Wage'})
        df_payroll['Hrs. Worked'] = (df_payroll['Min. Worked']/60).round(2)
        df_payroll = df_payroll.reindex(columns=['Name', 'Shift', 'Min. Worked', 'Hrs. Worked',  'Wage'])
        pay_period_sick_time = time_off.loc[time_off['Name']== name]['Sick Hrs'].sum()
        pay_period_vac_time = time_off.loc[time_off['Name']== name]['Vac Hrs'].sum()
        #paid vacation and sick leave
        if name in list(time_off['Name']):
            if pay_period_sick_time > 0:
                #this is hourly wage for sick cash out
                sick_amount = regular_rate
                df_sick = pd.DataFrame({'Name': name, 'Shift': ['Sick Leave Used'], 'Min. Worked': [60*pay_period_sick_time], 
                                        'Hrs. Worked':[pay_period_sick_time],
                                        'Wage': [sick_amount]})
                df_payroll= pd.concat([df_payroll, df_sick], ignore_index=True)
            if pay_period_vac_time > 0:
                vac_amount = regular_rate
                df_vac = pd.DataFrame({'Name': name, 'Shift': ['Vacation Payout'], 'Min. Worked': [60*pay_period_vac_time], 
                                       'Hrs. Worked':[pay_period_vac_time],
                                        'Wage': [vac_amount]})
                df_payroll= pd.concat([df_payroll, df_vac], ignore_index=True)
        #Add bonus
        if name in list(bonus['Name']):
            bonus_amount = bonus.loc[bonus['Name'] == name]['Bonus Amount'].sum()
            df_bonus = pd.DataFrame({'Name': name, 'Shift': ['Bonus'], 'Min. Worked': [60], 'Hrs. Worked':[1],
                                        'Wage': [bonus_amount]})
            df_payroll= pd.concat([df_payroll, df_bonus], ignore_index=True)
        #Add Premium Pay
        premium_hrs = bonus_df.loc[bonus_df.Name == name]['Premium Pay Hours'].sum()
        if premium_hrs > 0:
            df_premium = pd.DataFrame({'Name': name, 'Shift': ['Premium Pay'], 'Min. Worked': [premium_hrs*60], 
                                       'Hrs. Worked':[premium_hrs],
                                        'Wage': [(1.5*regular_rate).round(2)]})
            df_payroll= pd.concat([df_payroll, df_premium], ignore_index=True)    
        df_payroll['Gross Wages'] = df_payroll['Hrs. Worked'] * df_payroll['Wage']
        df_payroll=df_payroll.round(decimals=2)
        total_gross_wage = df_payroll['Gross Wages'].sum()
        #accrual
        accrued_df = deepcopy(accrued_hrs.loc[accrued_hrs.Staff == name])
        accrued_df['YTD Vacation Taken'] += pay_period_vac_time
        accrued_df['Sick Taken'] += pay_period_sick_time
        accrued_df['YTD Hours'] += total_hours_worked
        if accrued_df.Sub.iloc[0] == 0: #NON-SUB
            is_sub = ' '
            pay_period_accrued_vac = round(total_hours_worked * 0.04, 2)
            if accrued_df['YTD Vacation Accrued'].iloc[0] > 80:
                accrued_df['YTD Vacation Accrued'] = 80
        else: #SUB
            is_sub = 'SUB ONLY'
            pay_period_accrued_vac = 0 #subs do not accrue vacations
        accrued_df['YTD Vacation Accrued'] += pay_period_accrued_vac
        accrued_df['Vacation Balance'] = accrued_df['YTD Vacation Accrued'] + accrued_df['Vac. Hrs Carried Over'] - accrued_df['YTD Vacation Taken']
        accrued_df['Sick Balance'] = accrued_df['Sick Bank'] - accrued_df['Sick Taken']
        new_accrued_hrs = pd.concat([new_accrued_hrs, accrued_df], ignore_index=True)
        non_mgr_payroll.append({'header': pd.DataFrame(columns=[name, is_sub]), 
                                'summary': pd.DataFrame({ 
                                    'Total Gross Wage': total_gross_wage, 
                                    'Pay Period': PAY_PERIOD}, index=[0]), 
                                    'payroll': df_payroll,
                                    'accrued_A': pd.DataFrame({'Hrs. YTD': accrued_df['YTD Hours'].iloc[0], 
                                                                'Hrs. Worked This Period': total_hours_worked,
                                                                'Hire Date': staff_info.loc[staff_info.Name == name]['Hire Date'].iloc[0],
                                                                'Calendar Days Since Hire Date': staff_info.loc[staff_info.Name == name]['Days Elapsed Since Hire Date'].iloc[0]},
                                                                index=[0]).round(decimals=2),
                                    'accrued_B': pd.DataFrame({'Vac. Accrued YTD': accrued_df['YTD Vacation Accrued'].iloc[0], 
                                                                'Vac. Taken YTD': accrued_df['YTD Vacation Taken'].iloc[0], 
                                                                'Vac. Accrued This Period': pay_period_accrued_vac,
                                                                'Vac. Taken This Period': pay_period_vac_time,
                                                                'Vac. Hrs Carried Over': accrued_df['Vac. Hrs Carried Over'].iloc[0],
                                                                'Vac. Balance': accrued_df['Vacation Balance'].iloc[0]},index=[0]).round(decimals=2),
                                    'accrued_C': pd.DataFrame({'Sick Bank YTD': accrued_df['Sick Bank'].iloc[0], 
                                                                'Sick Taken YTD': accrued_df['Sick Taken'].iloc[0], 
                                                                'Sick Taken This Period':pay_period_sick_time,
                                                                'Sick Balance': accrued_df['Sick Balance'].iloc[0]},index=[0]).round(decimals=2)
                                    })
    return (non_mgr_payroll, new_accrued_hrs)

def manager_payroll(mgr, manager_rates, df_shift_merged, accrued_hrs, bonus_df, bonus, time_off, week_order, prepaid_last_time, PAY_PERIOD, PREPAY, new_accrued_hrs):
    '''
    Process payroll for managers and return manager payroll and acrrued hours.

    mgr -- list of managers
    manager_rates -- rates of managers
    df_shift_merged -- shift dataframe for the pay period
    accrued_hrs -- accrued hours (existing for manipulation)
    bonus_df -- bonus info as dataframe
    bonus -- bonus summarized by person
    time_off -- time taken off
    week_order -- order of weeks within the pay period
    prepaid_last_time -- list of shifts prepaid last time
    PAY_PERIOD -- a string of pay period
    PREPAY -- boolean, if we are prepaying in this period
    new_accrued_hrs -- new accrual dataframe
    '''
    mgr_payroll = []
    df_shift_merged = deepcopy(df_shift_merged)
    aggregations = {'Min. Worked': 'sum', 'Regular Hourly Wage': 'first', 'Name': 'first'}
    if len(prepaid_last_time) > 0:
        prepaid_ppl = prepaid_last_time['Name'].unique()
    else:
        prepaid_ppl = set()
    for name in mgr:
        df_indiv = df_shift_merged.loc[df_shift_merged['Name'] == name]
        df_indiv_worked = df_indiv[~df_indiv['Shift'].str.contains('-Not-Worked')]
        total_hours_worked = round(df_indiv_worked['Min. Worked'].sum()/60, 2)
        regular_rate = manager_rates.loc[manager_rates['Name'] == name]['Admin/Sick/Vacay Wage'].iloc[0]
        df_payroll = pd.DataFrame()
        if name == 'Mikayla Napier':
            MGR_weekly_salary = manager_rates.loc[manager_rates['Name'] == name]['Exempt Semi-monthly Salary'].iloc[0]
            df_payroll = pd.DataFrame({'Name': name, 'Shift': ['MGR Salary'], 'Min. Worked': [60], 'Regular Hourly Wage': [MGR_weekly_salary]})
        else:
            MGR_weekly_salary = manager_rates.loc[manager_rates['Name'] == name]['Exempt Weekly Salary'].iloc[0]
            df_weeks = split_by_work_week(df_indiv)
            for key in df_weeks.keys(): #each key is a timestamp
                df_weekly = deepcopy(df_weeks[key])
                df_payroll_weekly = pd.DataFrame()
                # add prepaid time
                if (key == week_order[0]) and (name in prepaid_ppl): #first week and prepaid
                    df_prepaid = pd.DataFrame({'Name': name, 'Shift': ['Prepaid Last Time'], 'Min. Worked': [60], 
                                               'Regular Hourly Wage': [-MGR_weekly_salary]})
                    df_payroll_weekly = pd.concat([df_payroll_weekly, df_prepaid], ignore_index=True)
                    df_weekly = pd.concat([prepaid_last_time.loc[prepaid_last_time.Name == name], df_weekly],ignore_index=True)
                weekly_hours_worked = round(df_weekly['Min. Worked'].sum()/60, 2)
                exempt_hours_worked = round(df_weekly.loc[df_weekly['Shift'] != 'MGR-Direct-Care']['Min. Worked'].sum()/60, 2)
                overtime_hours = max(0, weekly_hours_worked-40)
                # Exempt
                if exempt_hours_worked >= (weekly_hours_worked - exempt_hours_worked) or ((key==week_order[-1]) and PREPAY): 
                    MGR_weekly_salary = manager_rates.loc[manager_rates['Name'] == name]['Exempt Weekly Salary'].iloc[0]
                    tmp = pd.DataFrame({'Name': name, 'Shift': ['MGR Salary'], 'Min. Worked': [60], 'Regular Hourly Wage': [MGR_weekly_salary]})
                    df_payroll_weekly = pd.concat([df_payroll_weekly,tmp], ignore_index=True)
                else: #Non exempt
                    aggregations2 = {'Min. Worked': 'sum',  'Name': 'first'}
                    tmp = deepcopy(df_weekly[['Name', 'Shift', 'Min. Worked']])
                    tmp = tmp.groupby('Shift').agg(aggregations2)
                    tmp = tmp.reset_index() 
                    tmp['Regular Hourly Wage'] = [regular_rate]*len(tmp)
                    df_payroll_weekly = pd.concat([df_payroll_weekly,tmp], ignore_index=True)
                    overtime_hours = max(0, weekly_hours_worked-40)
                    if overtime_hours > 0:
                        BOT_pay_rate = (0.5 * regular_rate).round(2)
                        df_overtime = pd.DataFrame({'Name': name, 'Shift': [f'OT Extra Pay ({key})'], 
                                                    'Min. Worked': [overtime_hours], 'Regular Hourly Wage': [BOT_pay_rate]})
                        df_payroll_weekly= pd.concat([df_payroll_weekly, df_overtime], ignore_index=True)
                df_payroll=pd.concat([df_payroll, df_payroll_weekly], ignore_index=True)
        #add holiday bonus
        holiday_work_time = df_indiv_worked['Holiday Worked Duration (Minutes)'].sum()
        if holiday_work_time > 0:
            df_holiday_pay = pd.DataFrame({'Name': name, 'Shift': ['Holiday Extra Pay'], 'Min. Worked': [holiday_work_time], 
                                        'Regular Hourly Wage': [round(regular_rate/2, 2)]})
            df_payroll= pd.concat([df_payroll, df_holiday_pay], ignore_index=True)
        #add sick and vacation
        pay_period_sick_time = time_off.loc[time_off['Name']== name]['Sick Hrs'].sum()
        pay_period_vac_time = time_off.loc[time_off['Name']== name]['Vac Hrs'].sum()      
        if name in list(time_off['Name']):
            if pay_period_sick_time > 0:
                #this is hourly wage for sick cash out
                sick_amount = regular_rate
                df_sick = pd.DataFrame({'Name': name, 'Shift': ['Sick Leave Used'], 'Min. Worked': [60*pay_period_sick_time], 
                                        'Regular Hourly Wage': [sick_amount]})
                df_payroll= pd.concat([df_payroll, df_sick], ignore_index=True)
            if pay_period_vac_time > 0:
                vac_amount = regular_rate
                df_vac = pd.DataFrame({'Name': name, 'Shift': ['Vacation Payout'], 'Min. Worked': [60*pay_period_vac_time], 
                                       'Regular Hourly Wage': [vac_amount]})
                df_payroll= pd.concat([df_payroll, df_vac], ignore_index=True)
        #bonus
        if name in list(bonus['Name']):
            bonus_amount = bonus.loc[bonus['Name'] == name]['Bonus Amount'].sum()
            df_bonus = pd.DataFrame({'Name': name, 'Shift': ['Bonus'], 'Min. Worked': [60], 
                                     'Regular Hourly Wage': [bonus_amount]})
            #Add to payroll
            df_payroll= pd.concat([df_payroll, df_bonus], ignore_index=True)
        #premium
        premium_hrs = bonus_df.loc[bonus_df.Name == name]['Premium Pay Hours'].sum()
        if premium_hrs > 0:
            df_premium = pd.DataFrame({'Name': name, 'Shift': ['Premium Pay'], 'Min. Worked': [premium_hrs*60], 
                                       'Regular Hourly Wage': [1.5*regular_rate]})
            df_payroll= pd.concat([df_payroll, df_premium], ignore_index=True)
        df_payroll = df_payroll.groupby('Shift').agg(aggregations)
        df_payroll = df_payroll.reset_index()
        df_payroll = df_payroll.rename(columns={'Regular Hourly Wage': 'Wage'})
        df_payroll['Hrs. Worked'] = round(df_payroll['Min. Worked']/60, 2)
        df_payroll = df_payroll.reindex(columns=['Name', 'Shift', 'Min. Worked', 'Hrs. Worked',  'Wage'])
        df_payroll['Gross Wages'] = (df_payroll['Hrs. Worked'] * df_payroll['Wage']).round(2)
        total_gross_wage = df_payroll['Gross Wages'].sum()
        #Accrual
        accrued_df = deepcopy(accrued_hrs.loc[accrued_hrs.Staff == name])
        accrued_df['YTD Vacation Taken'] += pay_period_vac_time
        accrued_df['Sick Taken'] += pay_period_sick_time
        accrued_df['YTD Hours'] += total_hours_worked
        pay_period_accrued_vac = round(total_hours_worked*0.068,2)
        accrued_df['YTD Vacation Accrued'] += pay_period_accrued_vac
        if accrued_df['YTD Vacation Accrued'].iloc[0] > 136:
            accrued_df['YTD Vacation Accrued'] = 136
        accrued_df['Vacation Balance'] = accrued_df['YTD Vacation Accrued'] + accrued_df['Vac. Hrs Carried Over'] - accrued_df['YTD Vacation Taken']
        accrued_df['Sick Balance'] = accrued_df['Sick Bank'] - accrued_df['Sick Taken']
        new_accrued_hrs = pd.concat([new_accrued_hrs, accrued_df], ignore_index=True)
        is_sub = ' '
        mgr_payroll.append({'header': pd.DataFrame(columns=[name, is_sub]), 
                            'summary': pd.DataFrame({'Total Hours Worked': total_hours_worked, 
                                                    'Total Gross Wage': total_gross_wage, 
                                                    'Pay Period': PAY_PERIOD}, index=[0]).round(decimals=2), 
                            'payroll': df_payroll.round(decimals=2), 
                            'accrued_A': pd.DataFrame({'Hrs. YTD': accrued_df['YTD Hours'].iloc[0], 
                                                        'Hrs. Worked This Period': total_hours_worked,
                                                        'Hire Date': manager_rates.loc[manager_rates.Name == name]['Hire Date'].iloc[0],
                                                        'Calendar Days Since Hire Date': manager_rates.loc[manager_rates.Name == name]['Days Elapsed Since Hire Date'].iloc[0]},
                                                        index=[0]).round(decimals=2),
                            'accrued_B': pd.DataFrame({'Vac. Accrued YTD': accrued_df['YTD Vacation Accrued'].iloc[0], 
                                                        'Vac. Taken YTD': accrued_df['YTD Vacation Taken'].iloc[0], 
                                                        'Vac. Accrued This Period': pay_period_accrued_vac,
                                                        'Vac. Taken This Period': pay_period_vac_time,
                                                        'Vac. Hrs Carried Over': accrued_df['Vac. Hrs Carried Over'].iloc[0],
                                                        'Vac. Balance': accrued_df['Vacation Balance'].iloc[0]},index=[0]).round(decimals=2),
                            'accrued_C': pd.DataFrame({'Sick Bank YTD': accrued_df['Sick Bank'].iloc[0], 
                                                        'Sick Taken YTD': accrued_df['Sick Taken'].iloc[0], 
                                                        'Sick Taken This Period':pay_period_sick_time,
                                                        'Sick Balance': accrued_df['Sick Balance'].iloc[0]},index=[0]).round(decimals=2)
                            })      
    return (mgr_payroll, new_accrued_hrs)

def non_manager_weekly_breakdown(non_mgr, df_shift_merged, prepaid_last_time, week_order):  
    '''
    Generate weekly breakdown for non-managers
    ''' 
    #Create a list that stores the payroll dictionary
    non_mgr_payroll = []
    prepaid_last_time = deepcopy(prepaid_last_time)
    df_shift_merged = deepcopy(df_shift_merged)
    if len(prepaid_last_time) > 0:
        prepaid_ppl = prepaid_last_time['Name'].unique()
    else:
        prepaid_ppl = set()
    #For each non-manager
    for _, name in enumerate(non_mgr):
        #subset to the individual's shift
        aggregations = {'Min. Worked': 'sum', 'BOT Hourly Wage': 'first', 'Name': 'first'}
        df_indiv = df_shift_merged.loc[df_shift_merged['Name'] == name]
        df_weeks = split_by_work_week(df_indiv)
        for key in df_weeks.keys(): #each key is a timestamp
            df_weekly = deepcopy(df_weeks[key])
            df_payroll = deepcopy(df_weekly[['Shift', 'Min. Worked', 'BOT Hourly Wage', 'Name']])
            if (key == week_order[0]) and (name in prepaid_ppl): #first week and prepaid
                df_weekly = pd.concat([prepaid_last_time.loc[prepaid_last_time.Name == name], df_weekly], ignore_index=True)
                prepaid_concat = prepaid_last_time.loc[prepaid_last_time.Name == name]
                prepaid_concat['Shift'] = 'PREPAID ' + prepaid_concat['Shift']
                prepaid_concat=prepaid_concat[['Shift', 'Min. Worked', 'BOT Hourly Wage', 'Name']]
                df_payroll= pd.concat([df_payroll, prepaid_concat], ignore_index=True)
            df_weekly_worked = df_weekly[~df_weekly['Shift'].str.contains('-Not-Worked')]
            weekly_hours_worked = round(df_weekly_worked['Min. Worked'].sum()/60, 2)
            weekly_hours_paid = round(df_weekly['Min. Worked'].sum()/60, 2)
            overtime_hours = max(0, weekly_hours_worked - 40)
            ot_rate =((df_weekly['BOT Hourly Wage']*(df_weekly['Min. Worked']/60).round(2)).sum()/weekly_hours_paid)
            if overtime_hours > 0:
                df_overtime = pd.DataFrame({'Name': name, 'Shift': [f'OT Extra Pay ({key})'], 'Min. Worked': [round(overtime_hours*60, 2)], 
                                        'BOT Hourly Wage': [round(ot_rate/2, 2)]})
                #Add to payroll
                df_payroll= pd.concat([df_payroll, df_overtime], ignore_index=True)
            df_payroll = df_payroll.groupby('Shift').agg(aggregations)
            df_payroll = df_payroll.reset_index()
            #Dealing with holiday
            #reassign variables to remove prepaid hours
            df_weekly = deepcopy(df_weeks[key])
            df_weekly_worked = df_weekly[~df_weekly['Shift'].str.contains('-Not-Worked')]
            if df_weekly_worked['Holiday Worked Duration (Minutes)'].sum() > 0:
                df_holiday_pay = df_weekly_worked[['Name', 'Shift', 'Holiday Worked Duration (Minutes)', 'BOT Hourly Wage']]
                df_holiday_pay = df_holiday_pay[df_holiday_pay['Holiday Worked Duration (Minutes)'] != 0]
                df_holiday_pay['Shift'] = df_holiday_pay['Shift'].apply(lambda x: x + ' Holiday Extra Pay')
                df_holiday_pay['BOT Hourly Wage'] = df_holiday_pay['BOT Hourly Wage']*0.5
                df_holiday_pay = df_holiday_pay.rename(columns={'Holiday Worked Duration (Minutes)': 'Min. Worked'})
                df_holiday_pay = df_holiday_pay.groupby('Shift').agg(aggregations)
                df_holiday_pay = df_holiday_pay.reset_index()
                df_payroll= pd.concat([df_payroll, df_holiday_pay], ignore_index=True)
            df_payroll = df_payroll.rename(columns={'BOT Hourly Wage': 'Wage'})
            df_payroll['Hrs. Worked'] = (df_payroll['Min. Worked']/60).round(2)
            df_payroll['Hrs. Paid'] = df_payroll['Hrs. Worked']
            df_payroll = df_payroll.reindex(columns=['Name', 'Shift', 'Min. Worked', 'Hrs. Worked', 'Hrs. Paid', 'Wage'])
            df_payroll['Gross Wages'] = df_payroll['Hrs. Worked'] * df_payroll['Wage']
            #In Nov-Paid Hrs. We exclude IHSS asleep
            df_payroll['Nova-Paid Hrs.'] = df_payroll['Hrs. Worked']
            #Correct Hours worked
            df_payroll.loc[df_payroll['Shift'].str.contains('-Not-Worked'), 'Hrs. Worked'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Hrs. Paid'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Hrs. Worked'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('Asleep') & ~df_payroll['Shift'].str.contains('Holiday Extra Pay'), 'Nova-Paid Hrs.'] = 0
            df_payroll['Nova-Paid Gross Wages'] = df_payroll['Gross Wages']
            df_payroll.loc[df_payroll['Shift'].str.contains('Asleep') & ~df_payroll['Shift'].str.contains('Holiday Extra Pay'), 'Nova-Paid Gross Wages'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Gross Wages'] = 0
            column_order = ['Name', 'Shift', 'Hrs. Worked', 'Hrs. Paid', 'Nova-Paid Hrs.', 'Wage', 'Gross Wages', 'Nova-Paid Gross Wages']
            df_payroll = df_payroll[column_order]
            real_wages_paid = df_payroll['Nova-Paid Gross Wages'].sum() - df_payroll.loc[df_payroll['Shift'].str.contains('PREPAID')]['Nova-Paid Gross Wages'].sum()
            df_sum = pd.DataFrame(df_payroll.round(decimals=2).sum(axis=0)).T
            df_sum['Name']="TOTAL"
            df_sum['Shift']="---"
            overtime_rate = round((df_sum['Gross Wages']/df_sum['Hrs. Paid']).iloc[0], 2)
            df_payroll = pd.concat([df_payroll, df_sum], ignore_index=True).append(pd.DataFrame(index=[1]))
            non_mgr_payroll.append({'header': pd.DataFrame(columns=[name, 'Week of ' + key]), 
                                    'summary': pd.DataFrame({ 
                                        'Weekly Nova-Paid Gross Wages - Prepaid Wages': real_wages_paid, 
                                        'Total Gross Wages / Total Hrs. Paid':overtime_rate}, index=[0]).round(decimals=2), 
                                        'payroll': df_payroll.round(decimals=2)
                                        })
    return non_mgr_payroll

def manager_weekly_breakdown(mgr, manager_rates, df_shift_merged, week_order, prepaid_last_time, PAY_PERIOD, PREPAY):
    '''
    Generate the weekly breakdown for managers
    '''
    mgr_payroll = []
    df_shift_merged = deepcopy(df_shift_merged)
    if len(prepaid_last_time) > 0:
        prepaid_ppl = prepaid_last_time['Name'].unique()
    else:
        prepaid_ppl = set()
    #aggregations = {'Min. Worked': 'sum', 'Regular Hourly Wage': 'first', 'Name': 'first'}
    for name in mgr:
        df_indiv = df_shift_merged.loc[df_shift_merged['Name'] == name]
        df_indiv_worked = df_indiv[~df_indiv['Shift'].str.contains('-Not-Worked')]
        df_payroll = pd.DataFrame(columns=['Name', 'Shift','Min. Worked', 'Regular Hourly Wage'])
        holiday_work_time = df_indiv_worked['Holiday Worked Duration (Minutes)'].sum()
        non_exempt_rate = manager_rates.loc[manager_rates['Name'] == name]['Non-exempt Hourly Wage'].iloc[0]
        #holiday extra pay
        if holiday_work_time > 0:
            df_holiday_pay = pd.DataFrame({'Name': name, 'Shift': ['Holiday Extra Pay'], 'Min. Worked': [holiday_work_time], 
                                            'Regular Hourly Wage': [round(0.5 * non_exempt_rate, 2)]})
            df_payroll = df_holiday_pay
        if name == 'Mikayla Napier':
            MGR_weekly_salary = manager_rates.loc[manager_rates['Name'] == name]['Exempt Semi-monthly Salary'].iloc[0]
            df_payroll = pd.concat([df_payroll, pd.DataFrame({'Name': name, 'Shift': ['MGR Salary'], 'Min. Worked': [60], 'Regular Hourly Wage': [MGR_weekly_salary]})])
            key = PAY_PERIOD
            df_payroll['Hrs. Worked'] = round(df_payroll['Min. Worked']/60, 2)
            df_payroll['Hrs. Paid'] = df_payroll['Hrs. Worked']
            df_payroll = df_payroll.rename(columns={'Regular Hourly Wage': 'Wage'})
            real_wages_paid = round((df_payroll['Wage'] * df_payroll['Hrs. Worked']).sum(), 2)
            df_payroll['Gross Wages'] = df_payroll['Hrs. Worked'] * df_payroll['Wage']
            df_payroll['Nova-Paid Hrs.'] = df_payroll['Hrs. Worked']
            #Correct Hours worked
            df_payroll.loc[df_payroll['Shift'].str.contains('-Not-Worked'), 'Hrs. Worked'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('Asleep') & ~df_payroll['Shift'].str.contains('Holiday Extra Pay'), 'Nova-Paid Hrs.'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Hrs. Paid'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Hrs. Worked'] = 0
            df_payroll['Nova-Paid Gross Wages'] = df_payroll['Gross Wages']
            df_payroll.loc[df_payroll['Shift'].str.contains('Asleep') & ~df_payroll['Shift'].str.contains('Holiday Extra Pay'), 'Nova-Paid Gross Wages'] = 0
            df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Gross Wages'] = 0
            column_order = ['Name', 'Shift', 'Hrs. Worked', 'Hrs. Paid', 'Nova-Paid Hrs.', 'Wage', 'Gross Wages', 'Nova-Paid Gross Wages']
            df_payroll = df_payroll[column_order]
            df_sum = pd.DataFrame(df_payroll.round(decimals=2).sum(axis=0)).T
            df_sum['Name']="TOTAL"
            df_sum['Shift']="---"
            df_payroll = pd.concat([df_payroll, df_sum], ignore_index=True).append(pd.DataFrame(index=[1]))
            mgr_payroll.append({'header': pd.DataFrame(columns=[name, 'Weeks of ' + key]), 
                                        'summary': pd.DataFrame({ 
                                            'Weekly Nova-Paid Gross Wages - Prepaid Wages': real_wages_paid
                                                                }, index=[0]).round(decimals=2), 
                                            'payroll': df_payroll.round(decimals=2)
                                            }) 
        else:
            MGR_weekly_salary = manager_rates.loc[manager_rates['Name'] == name]['Exempt Weekly Salary'].iloc[0]
            df_weeks = split_by_work_week(df_indiv)
            for key in df_weeks.keys(): #each key is a timestamp
                df_weekly = deepcopy(df_weeks[key])
                df_weekly_worked = df_weekly[~df_weekly['Shift'].str.contains('-Not-Worked')]
                df_payroll = pd.DataFrame()
                if (key==week_order[0]) and (name in prepaid_ppl): #prepaid
                    df_prepaid = pd.DataFrame({'Name': name, 'Shift': ['PREPAID MGR Salary'], 'Min. Worked': [60], 
                                               'Regular Hourly Wage': [MGR_weekly_salary]})
                    df_payroll = pd.concat([df_payroll, df_prepaid], ignore_index=True)
                    df_weekly = pd.concat([df_weekly, prepaid_last_time], ignore_index=True)
                    df_weekly_worked = df_weekly[~df_weekly['Shift'].str.contains('-Not-Worked')]
                weekly_hours_worked = (df_weekly_worked['Min. Worked'].sum()/60).round(2)
                exempt_hours_worked = (df_weekly_worked.loc[df_weekly_worked['Shift'] != 'MGR-Direct-Care']['Min. Worked'].sum()/60).round(2)
                overtime_hours = max(0, weekly_hours_worked-40)
                if exempt_hours_worked >= (weekly_hours_worked - exempt_hours_worked) or ((key==week_order[-1]) and PREPAY): # Exempt
                    MGR_weekly_salary = manager_rates.loc[manager_rates['Name'] == name]['Exempt Weekly Salary'].iloc[0]
                    tmp = pd.DataFrame({'Name': name, 'Shift': ['MGR Salary'], 'Min. Worked': [60], 'Regular Hourly Wage': [MGR_weekly_salary]})
                    df_payroll = pd.concat([df_payroll, tmp], ignore_index=True)
                else: # Non-exempt
                    aggregations2 = {'Min. Worked': 'sum',  'Name': 'first'}
                    tmp = deepcopy(df_weekly[['Name', 'Shift', 'Min. Worked']])
                    tmp = tmp.groupby('Shift').agg(aggregations2)
                    tmp = tmp.reset_index() 
                    tmp['Regular Hourly Wage'] = [non_exempt_rate]*len(tmp)
                    df_payroll = pd.concat([df_payroll, tmp], ignore_index=True)
                    overtime_hours = max(0, weekly_hours_worked-40)
                    if overtime_hours > 0:
                        BOT_pay_rate = 0.5 * non_exempt_rate
                        df_overtime = pd.DataFrame({'Name': name, 'Shift': [f'OT Extra Pay ({key})'], 'Min. Worked': [overtime_hours], 
                                                    'Regular Hourly Wage': [BOT_pay_rate]})
                        df_payroll= pd.concat([df_payroll, df_overtime], ignore_index=True)
                df_payroll = df_payroll.rename(columns={'Regular Hourly Wage': 'Wage'})
                df_payroll['Hrs. Worked'] = (df_payroll['Min. Worked']/60).round(2)
                df_payroll['Hrs. Paid'] = df_payroll['Hrs. Worked']
                df_payroll = df_payroll.reindex(columns=['Name', 'Shift', 'Min. Worked', 'Hrs. Worked', 'Hrs. Paid', 'Wage'])
                df_payroll['Gross Wages'] = df_payroll['Hrs. Worked'] * df_payroll['Wage']
                df_payroll['Nova-Paid Hrs.'] = df_payroll['Hrs. Worked']
                #Correct Hours worked
                df_payroll.loc[df_payroll['Shift'].str.contains('-Not-Worked'), 'Hrs. Worked'] = 0
                df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Hrs. Paid'] = 0
                df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Hrs. Worked'] = 0
                df_payroll.loc[df_payroll['Shift'].str.contains('Asleep') & ~df_payroll['Shift'].str.contains('Holiday Extra Pay'), 'Nova-Paid Hrs.'] = 0
                df_payroll['Nova-Paid Gross Wages'] = df_payroll['Gross Wages']
                df_payroll.loc[df_payroll['Shift'].str.contains('Asleep') & ~df_payroll['Shift'].str.contains('Holiday Extra Pay'), 'Nova-Paid Gross Wages'] = 0
                df_payroll.loc[df_payroll['Shift'].str.contains('OT Extra'), 'Gross Wages'] = 0
                column_order = ['Name', 'Shift', 'Hrs. Worked', 'Hrs. Paid', 'Nova-Paid Hrs.', 'Wage', 'Gross Wages', 'Nova-Paid Gross Wages']
                df_payroll = df_payroll[column_order]
                df_sum = pd.DataFrame(df_payroll.loc[~df_payroll['Shift'].str.contains('PREPAID')].round(decimals=2).sum(axis=0)).T
                df_sum['Name']="TOTAL"
                df_sum['Shift']="---"
                #when managers are prepaid they are prepaid for the whole week
                real_wages_paid = df_payroll['Nova-Paid Gross Wages'].sum() - 2*df_payroll.loc[df_payroll['Shift'].str.contains('PREPAID')]['Nova-Paid Gross Wages'].sum()
                df_payroll = pd.concat([df_payroll, df_sum], ignore_index=True).append(pd.DataFrame(index=[1]))
                mgr_payroll.append({'header': pd.DataFrame(columns=[name, 'Week of ' + key]), 
                                        'summary': pd.DataFrame({ 
                                            'Weekly Nova-Paid Gross Wages - Prepaid Wages': real_wages_paid
                                                                }, index=[0]).round(decimals=2), 
                                            'payroll': df_payroll.round(decimals=2)
                                            })   
    return mgr_payroll

def generate_payroll(df_shift_merged, accrued_hrs, bonus_df, bonus, time_off, manager_rates, staff_info, prepaid_last_time, 
                     PAY_PERIOD, week_order, PREPAY):
    '''
    Generate payroll files for human readers. This is a wrapper for four methods:
    
    non_manager_payroll()
    manager_payroll()
    non_manager_weekly_breakdown()
    manager_weekly_breakdown()

    Return results from the four methods
    '''
    staff_names = set().union(*[df_shift_merged.Name, bonus.Name, time_off.Name]).intersection(set().union(*[staff_info.Name, manager_rates.Name]))
    manager_status = [is_manager(i,manager_rates) for i in staff_names]
    non_mgr = [] #list of names
    mgr = [] #list of names
    for i, name in enumerate(staff_names):
        if manager_status[i]:
            mgr.append(name)
        else:
            non_mgr.append(name)
    new_accrued_hrs = deepcopy(accrued_hrs[~accrued_hrs['Staff'].isin(staff_names)]).round(2)
    df_shift_merged = df_shift_merged.round(2)
    accrued_hrs = accrued_hrs.round(2)
    bonus_df = bonus_df.round(2)
    bonus = bonus.round(2)
    time_off = time_off.round(2)
    staff_info = staff_info.round(2)
    prepaid_last_time = prepaid_last_time.round(2)
    non_mgr_pr = {}
    mgr_pr = {}
    non_mgr_bkd = {}
    mgr_bkd = {}
    if any(is_manager(name, manager_rates) == False for name in staff_names):
        non_mgr_pr, new_accrued_hrs = non_manager_payroll(non_mgr, df_shift_merged, accrued_hrs, bonus_df, bonus, time_off, staff_info,
                                                        week_order, prepaid_last_time, PAY_PERIOD, new_accrued_hrs)
        non_mgr_bkd = non_manager_weekly_breakdown(non_mgr, df_shift_merged, prepaid_last_time, week_order)
    if any(is_manager(name, manager_rates) for name in staff_names):
        mgr_pr, new_accrued_hrs = manager_payroll(mgr, manager_rates, df_shift_merged, accrued_hrs, bonus_df, bonus, time_off, week_order, 
                                              prepaid_last_time, PAY_PERIOD, PREPAY, new_accrued_hrs)
        mgr_bkd = manager_weekly_breakdown(mgr, manager_rates, df_shift_merged, week_order, prepaid_last_time, PAY_PERIOD, PREPAY)

    return (non_mgr_pr, mgr_pr, non_mgr_bkd, mgr_bkd, new_accrued_hrs)

def output_payroll_files(save_path, df_shift_merged, staff_info, non_mgr_pr, mgr_pr, non_mgr_bkd, mgr_bkd, new_accrued_hrs, original_bonus_df, time_off_as_shifts, non_manager_rates, manager_rates, prepaid_hours, df_after_pay_period, PAY_PERIOD):
    '''
    Output payroll files and save to an excel.
    '''
    df_shift_merged['Holiday Worked Duration (Hours)'] = (df_shift_merged['Holiday Worked Duration (Minutes)']/60).round(2)
    df_shift_merged['Hrs. Worked'] = df_shift_merged['Min. Worked']/60
    df_shift_merged = df_shift_merged.round(decimals=2)
    df_shift_merged['Day of the Week'] = df_shift_merged['CIDT'].dt.day_name()
    df_shift_merged = df_shift_merged[['Name', 'First Name', 'Last Name', 'Shift_original', 'Shift','Day of the Week', 'Check-In Date', 'Check-In Time', 
                                       'Check-Out Date', 'Check-Out Time', 'Min. Worked', 'Hrs. Worked',  'Regular Hourly Wage', 'BOT Hourly Wage', 'Accrual Rate',
                                       'Error Check', 'CIDT', 'CODT','Holiday Worked Duration (Minutes)','Holiday Worked Duration (Hours)']]
    df_shift_merged = df_shift_merged.reindex(columns=['Name', 'First Name', 'Last Name', 'Shift_original', 'Shift','Day of the Week', 'Check-In Date', 
                                                       'Check-In Time', 'Check-Out Date', 'Check-Out Time', 'Min. Worked', 'Hrs. Worked',  'Regular Hourly Wage', 
                                                       'BOT Hourly Wage', 'Accrual Rate','Error Check', 'CIDT', 'CODT','Holiday Worked Duration (Minutes)',
                                                       'Holiday Worked Duration (Hours)'])
    
    payroll_list = [ *non_mgr_pr, *mgr_pr ]
    sorted_payroll_list = sorted(payroll_list, key=lambda x: x['header'].columns[0].split()[-1])
    bkd_list = [ *non_mgr_bkd, *mgr_bkd ]
    sorted_bkd_list = sorted(bkd_list, key=lambda x: x['header'].columns[0].split()[-1])
    #Output payroll
    payroll_path = save_path+"/"+f"PAYROLL OUTPUT - {PAY_PERIOD}.xlsx"
    writer = pd.ExcelWriter(payroll_path) 
    startrow = 0

    for person in sorted_payroll_list:
        for df in [person['header'], person['payroll'],person['summary'], person['accrued_A'],person['accrued_B'],person['accrued_C']]:
            df.to_excel(writer, engine="xlsxwriter",sheet_name='FINAL PAYROLL', startrow=startrow, index=False)
            startrow += (df.shape[0] + 1)
        startrow += 3
    writer.sheets['FINAL PAYROLL'].set_column('A:F', 24)

    startrow = 0
    name = "NOVA"
    for index, person in enumerate(sorted_bkd_list):
        new_name = person['header'].columns[0]
        if name.lower() != new_name.lower():
            if startrow!=0:
                startrow += 3
            person['header'].columns=person['header'].columns.str.upper()
        name = person['header'].columns[0]
        last_name = sorted_bkd_list[-1]
        for df in [person['header'], person['payroll'], person['summary']]:
            df.to_excel(writer, engine="xlsxwriter",sheet_name='WEEKLY BREAKDOWNS', startrow=startrow, index=False)
            startrow += (df.shape[0] + 1)
        startrow += 2
    writer.sheets['WEEKLY BREAKDOWNS'].set_column('A:H', 40)

    df_shift_merged = pd.concat([df_shift_merged, time_off_as_shifts], ignore_index=True)
    df_shift_merged = df_shift_merged.sort_values(by=['Last Name', 'CIDT'])
    df_shift_merged.to_excel(writer, sheet_name="SHIFT BREAKDOWNS", index=False)

    for column in df_shift_merged:
        column_length = max(df_shift_merged[column].astype(str).map(len).max(), len(column))
        col_idx =df_shift_merged.columns.get_loc(column)
        writer.sheets['SHIFT BREAKDOWNS'].set_column(col_idx, col_idx, column_length)

    writer.save() 
    #output tracker
    new_tracker_path = save_path+"/"+f"NEW TRACKER - {PAY_PERIOD}.xlsx"
    writer = pd.ExcelWriter(new_tracker_path) 
    new_accrued_hrs = new_accrued_hrs.sort_values(by='Staff', key=lambda x: x.str.split().str[-1])
    staff_info = staff_info.sort_values(by='Name', key=lambda x: x.str.split().str[-1])
    columns_to_keep = ['Full Name', 'First Name', 'Last Name']
    for column in original_bonus_df.columns:
        if column not in columns_to_keep:
            original_bonus_df[column] = [np.nan]*len(original_bonus_df)

    original_bonus_df.to_excel(writer, sheet_name='NEW PTO & BONUS INFO', index=False)
    non_manager_rates.to_excel(writer, sheet_name='SHIFT INFO', index=False)
    staff_info.to_excel(writer, sheet_name='STAFF INFO', index=False)
    manager_rates.to_excel(writer, sheet_name='MANAGER INFO', index=False)
    new_accrued_hrs.to_excel(writer, sheet_name="HRS & ACCRUALS", index=False)
    prepaid_hours.to_excel(writer, sheet_name='IGNORE! (Prepaid Shifts)', index=False)
    df_after_pay_period.to_excel(writer, sheet_name='IGNORE! (Next Period Shifts)', index=False)

    for column in original_bonus_df:
        column_length = max(original_bonus_df[column].astype(str).map(len).max(), len(column))
        col_idx = original_bonus_df.columns.get_loc(column)
        writer.sheets['NEW PTO & BONUS INFO'].set_column(col_idx, col_idx, column_length)

    for column in non_manager_rates:
        column_length = max(non_manager_rates[column].astype(str).map(len).max(), len(column))
        col_idx = non_manager_rates.columns.get_loc(column)
        writer.sheets['SHIFT INFO'].set_column(col_idx, col_idx, column_length)

    for column in new_accrued_hrs:
        column_length = max(new_accrued_hrs[column].astype(str).map(len).max(), len(column))
        col_idx =new_accrued_hrs.columns.get_loc(column)
        writer.sheets['HRS & ACCRUALS'].set_column(col_idx, col_idx, column_length)

    for column in manager_rates:
        column_length = max(manager_rates[column].astype(str).map(len).max(), len(column))
        col_idx = manager_rates.columns.get_loc(column)
        writer.sheets['MANAGER INFO'].set_column(col_idx, col_idx, column_length)

    for column in staff_info:
        column_length = max(staff_info[column].astype(str).map(len).max(), len(column))
        col_idx = staff_info.columns.get_loc(column)
        writer.sheets['STAFF INFO'].set_column(col_idx, col_idx, column_length)

    writer.save() 

    # Load the Excel file
    workbook = load_workbook(new_tracker_path)

    # Select the active sheet
    worksheet = workbook['HRS & ACCRUALS']

    # Apply the background color to columns A to G
    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for column in 'ABCDEFG':
        for cell in worksheet[column]:
            cell.fill = fill

    # Save the modified Excel file
    workbook.save(new_tracker_path)

def output_payroll_for_one(selected_name, save_path, df_shift_merged, non_mgr_pr, mgr_pr, non_mgr_bkd, mgr_bkd, time_off_as_shifts, PAY_PERIOD):
    '''
    Output payroll for just one person.
    '''
    df_shift_merged['Holiday Worked Duration (Hours)'] = (df_shift_merged['Holiday Worked Duration (Minutes)']/60).round(2)
    df_shift_merged['Hrs. Worked'] = df_shift_merged['Min. Worked']/60
    df_shift_merged = df_shift_merged.round(decimals=2)
    df_shift_merged['Day of the Week'] = df_shift_merged['CIDT'].dt.day_name()
    df_shift_merged = df_shift_merged[['Name', 'First Name', 'Last Name', 'Shift_original', 'Shift','Day of the Week', 'Check-In Date', 'Check-In Time', 
                                    'Check-Out Date', 'Check-Out Time', 'Min. Worked', 'Hrs. Worked',  'Regular Hourly Wage', 'BOT Hourly Wage', 'Accrual Rate',
                                    'Error Check', 'CIDT', 'CODT','Holiday Worked Duration (Minutes)','Holiday Worked Duration (Hours)']]
    df_shift_merged = df_shift_merged.reindex(columns=['Name', 'First Name', 'Last Name', 'Shift_original', 'Shift','Day of the Week', 'Check-In Date', 
                                                    'Check-In Time', 'Check-Out Date', 'Check-Out Time', 'Min. Worked', 'Hrs. Worked',  'Regular Hourly Wage', 
                                                    'BOT Hourly Wage', 'Accrual Rate','Error Check', 'CIDT', 'CODT','Holiday Worked Duration (Minutes)',
                                                    'Holiday Worked Duration (Hours)'])
    payroll_list = [ *non_mgr_pr, *mgr_pr ]
    bkd_list = [ *non_mgr_bkd, *mgr_bkd ]
    sorted_bkd_list = sorted(bkd_list, key=lambda x: x['header'].columns[0].split()[-1])
    #Output payroll
    payroll_path = save_path+"/"+f"OFF CYCLE PAYROLL OUTPUT - {selected_name} - {PAY_PERIOD}.xlsx"
    writer = pd.ExcelWriter(payroll_path) 
    startrow = 0
    person  = [i for i in payroll_list if list(i['header'])[0] == selected_name][0]
    for df in [person['header'], person['payroll'],person['summary'], person['accrued_A'],person['accrued_B'],person['accrued_C']]:
        df.to_excel(writer, engine="xlsxwriter",sheet_name='FINAL PAYROLL', startrow=startrow, index=False)
        startrow += (df.shape[0] + 1)
    writer.sheets['FINAL PAYROLL'].set_column('A:F', 24)
    startrow = 0
    name = "NOVA"
    for index, person in enumerate(sorted_bkd_list):
        new_name = person['header'].columns[0]
        if name.lower() != new_name.lower():
            if startrow!=0:
                startrow += 3
            person['header'].columns=person['header'].columns.str.upper()
        name = person['header'].columns[0]
        last_name = sorted_bkd_list[-1]
        for df in [person['header'], person['payroll'], person['summary']]:
            df.to_excel(writer, engine="xlsxwriter",sheet_name='WEEKLY BREAKDOWNS', startrow=startrow, index=False)
            startrow += (df.shape[0] + 1)
        startrow += 2
    try:
        writer.sheets['WEEKLY BREAKDOWNS'].set_column('A:H', 40)
    except:
        pass
    df_shift_merged = pd.concat([df_shift_merged, time_off_as_shifts], ignore_index=True)
    df_shift_merged = df_shift_merged.sort_values(by=['Last Name', 'CIDT'])
    df_shift_merged.to_excel(writer, sheet_name="SHIFT BREAKDOWNS", index=False)
    for column in df_shift_merged:
        column_length = max(df_shift_merged[column].astype(str).map(len).max(), len(column))
        col_idx =df_shift_merged.columns.get_loc(column)
        writer.sheets['SHIFT BREAKDOWNS'].set_column(col_idx, col_idx, column_length)
    writer.save()

def generate_invoice(df_shift_merged, manager_rates, non_manager_rates, staff_info, non_mgr_pr, mgr_pr):
    '''
    Genrate the Invoice using payroll files and save as an excel
    '''
    # bill_rates: map name of shift to the billing rate
    # other_shifts: list of non-billable shifts
    # shift_list: list of all shifts, both billable and non-billable

    bill_rates = {}
    other_shifts = []
    shift_list = []
    payroll_list = [ *non_mgr_pr, *mgr_pr ]
    for _, row in non_manager_rates.iterrows():
        # shift: name of the shift
        shift = row[0]
        shift_list.append(shift)
        # obtain billable shifts
        if not pd.isna(row[4]) and (isinstance(row[4], float) or isinstance(row[4], int)):
            bill_rates[shift] = row[4]
        # unbillable shifts
        else:
            other_shifts.append(shift)
    #print(bill_rates)
    #print(other_shifts)
    #print(shift_list)
    # map each person to their HSS level
    staff_hss = {}
    for _, row in staff_info.iterrows():
        name = row[0]
        hss_lvl = row["HSS Level"]
        # handles rare case where hss_lvl is nan, where we set to HSS1
        if not isinstance(hss_lvl, str):
            hss_lvl = "HSS1"
        staff_hss[name] = hss_lvl
    # map each person to ordered pair (shift, hours).
    # this stores the data on what shift each person did and 
    #   number of hours worked in that shift.
    # append integer to distinguish shifts by same person.
    #   e.g. Joe Shmoe0 is the first shift Joe Shmoe did;
    #   Joe Shmoe1 is the second shift he did, etc.
    payroll_dict = {}
    for employee in payroll_list:
        payroll = employee["payroll"]
        # employee did not work
        if payroll.empty:
            continue
        # row[0] = name, row[1] = shift, row[2] = hours
        for _, row in payroll.iterrows():
            payroll_dict[row[0] + str(_)] = (row[1], row[3])
    # for each value of the dictionary, multiply hours by rate
    # add result to the dictionary "output"
    # output maps shift code to [original gross hours, rate, billable, BST hours to insurance, BST hours to SARC]
    #   last two entries added in next code block, and billable is to be updated for BST (only count those to SARC)
    output = {}

    for _ in payroll_dict:
        curr_name = _[:-1]
        (curr_shift, curr_hours) = payroll_dict[_]
        split_shift = curr_shift.split("-")
        # name of shift to be displayed on invoice
        output_shift = curr_shift
        if curr_shift not in bill_rates and curr_shift != "CCR-Worked":
            continue
        if curr_shift == "CCR-Worked":
            curr_shift = staff_hss[curr_name]
            output_shift = curr_shift
        elif curr_shift != "CCR-Not-Worked":
            if len(split_shift) >= 2:
                if split_shift[-1] == "Worked":
                    if split_shift[-2] == "Not":
                        # cut off "-Not-Worked"
                        output_shift = curr_shift[:-11]
                    else:
                        # cut off "-Worked"
                        output_shift = curr_shift[:-7]
        # CCR-Not-Worked --> CCR
        else:
            output_shift = "CCR"
        # map to (hours, bill_rate, amount)
        # initialize if entry not already present
        if output_shift not in output:
            output[output_shift] = [0, bill_rates[curr_shift], 0]
        output[output_shift][0] += curr_hours
    # round total hours (correct floating point error)
    # compute billable by multiplying hours with billing rate
    for _ in output:
        output[_][0] = round(output[_][0], 2)
        output[_][2] = round(output[_][0] * output[_][1], 2)

    # manually obtain information on BCBA through df_shift_merged
    # BCBA to SARC: "BCBA"
    # BCBA to BlueShield
    BCBA_BlueShield = ["Adaptive-Behavior-Treatment", "Family-Adaptive-Behavior-Treatment", "Report-Writing"]
    BCBA_hrs = 0
    BCBA_BlueShield_hrs = 0
    for index, row in df_shift_merged.iterrows():
        if row["Shift_original"] == "BCBA":
            BCBA_hrs += row["Hrs. Worked"]
        elif row["Shift_original"] in BCBA_BlueShield:
            BCBA_hrs += row["Hrs. Worked"]
            BCBA_BlueShield_hrs += row["Hrs. Worked"]

    output["BCBA"] = [round(BCBA_hrs, 2), bill_rates["BCBA"], round(BCBA_hrs * bill_rates["BCBA"], 2)]
    ####################################################################################################
    # col 5 (E): Original Gross Hours (already there)
    # hours = [_[0] for _ in output.values()]
    # col 6 (F): BST/BCBA Hours Billed to Insurance
    #   # RBT hours paid as BST1, # RBT hours paid as BST2, etc.
    #   if Shift_original is RBT
    # col 7 (G): Hours Billed to SARC
    #   deduct col 6 from original gross hours
    # hours = [_[0] for _ in output.values()] (need to modify!!!)
    # col 8 (H): shift rates (already there)
    # rates = [_[1] for _ in output.values]
    # col 9 (I): billed amounts; change for BST, it is (col 7) * (col 8)
    # billable = [_[2] for _ in output.values()]
    # [hours, rates, billable, BST_insurance_hrs, BST_SARC_hrs]
    ####################################################################################################
    # col 6
    # RBT_dict: maps each BST to number of RBT hours
    RBT_dict = {}
    df_RBT = df_shift_merged[df_shift_merged["Shift_original"] == "RBT"]
    for index, row in df_RBT.iterrows():
        curr_shift = row["Shift"]
        if curr_shift not in RBT_dict:
            RBT_dict[curr_shift] = 0
        RBT_dict[curr_shift] += row["Hrs. Worked"]
    # add hours billed to insurance and BST hours to SARC
    for shift in output:
        if shift in RBT_dict:
            RBT_hours = RBT_dict[shift]
            shift_info = output[shift]
            # add col 6 info
            shift_info.append(round(RBT_hours, 2))
            BST_SARC = shift_info[0] - shift_info[3]
            if abs(BST_SARC) <= 0.01:
                BST_SARC = 0
                shift_info[3] = shift_info[0]
            # add col 7 info
            shift_info.append(round(BST_SARC, 2))
            # update col 9 info
            shift_info[2] = round(shift_info[4] * shift_info[1], 2)
            output[shift] = shift_info
        elif shift == "BCBA":
            shift_info = output[shift]
            shift_info.append(round(BCBA_BlueShield_hrs, 2))
            BCBA_SARC = shift_info[0] - shift_info[3]
            if abs(BCBA_SARC) <= 0.01:
                BCBA_SARC = 0
                shift_info[3] = shift_info[0]
            shift_info.append(round(BCBA_SARC, 2))
            shift_info[2] = round(shift_info[4] * shift_info[1], 2)
            output[shift] = shift_info
        else:
            output[shift].append(0)
            output[shift].append(output[shift][0])
    # list(range(8, len(manager_rates.columns))): list of indices
    #   of benefits columns
    benefits_cols = list(range(8, len(manager_rates.columns)))
    mgr_benefits = manager_rates.iloc[:, [0] + benefits_cols]
    mgr_benefits = mgr_benefits.transpose()
    mgr_benefits.columns = mgr_benefits.iloc[0]
    mgr_benefits = mgr_benefits.iloc[1:]
    mgr_benefits = mgr_benefits.reset_index()
    mgr_benefits = mgr_benefits.rename(columns={"index": "Benefit Name"})
    mgr_benefits = mgr_benefits.fillna(0)
    # add row to benefits indicating gross wage
    # 1) map each manager to total gross wage for the pay period
    mgr_gross = {}
    # managers are the last entries of payroll_list
    for _ in range(len(manager_rates.index)):
        name = payroll_list[-_-1]["payroll"].iloc[0][0]
        wage = payroll_list[-_-1]["summary"].iloc[0][1]
        mgr_gross[name] = wage
    # 2) construct row; make sure order is right
    wages = ["Wages"]
    for _, col in enumerate(mgr_benefits):
        if _ < 1:
            continue
        wages.append(mgr_gross[col])
    # 3) add row
    mgr_benefits.loc[len(mgr_benefits.index)] = wages
    # sum the columns
    totals = mgr_benefits.sum()
    totals = totals.to_frame().transpose()
    totals.iloc[0][0] = "Total"
    mgr_benefits = pd.concat([mgr_benefits, totals])
    df_benefits = mgr_benefits
    # obtain grand total quantity
    total_mgr = mgr_benefits.iloc[len(mgr_benefits.index)-1, 1:].sum()
    return (shift_list, output, mgr_benefits, df_benefits, total_mgr)

def output_invoice(save_path, shift_list, output, mgr_benefits, df_benefits, total_mgr, df_shift_merged, PAY_PERIOD):
    '''
    Output the invoice and return the underlying dataset
    '''
    save_path = save_path+"/"+f"INVOICE - {PAY_PERIOD}.xlsx"
    # custom sort
    for i in range(len(shift_list)):
        shift_split = shift_list[i].split("-")
        if len(shift_split) >= 2:
            if shift_split[-1] == "Worked":
                if shift_split[-2] != "Not":
                    shift_list[i] = shift_list[i][:-7]

    #print(shift_list)

    cat_size_order = CategoricalDtype(shift_list + ["MGR Benefits"],
                                    ordered=True)

    ####################################################################################################
    # col 5 (E): Original Gross Hours (already there)
    # shifts = output.keys()
    # col 6 (F): BST/BCBA Hours Billed to Insurance
    #   # RBT hours paid as BST1, # RBT hours paid as BST2, etc.
    # col 7 (G): Hours Billed to SARC
    #   deduct col 6 from original gross hours
    # hours = [_[0] for _ in output.values()] (need to modify)
    # col 8 (H): shift rates (already there)
    # rates = [_[1] for _ in output.values]
    # col 9 (I): billed amounts; change for BST, it is (col 7) * (col 8)
    # billable = [_[2] for _ in output.values()]
    ####################################################################################################

    shifts = output.keys()
    hours = [_[0] for _ in output.values()]
    rates = [_[1] for _ in output.values()]
    billable = [_[2] for _ in output.values()]
    BST_insurance_hrs = [_[3] for _ in output.values()]
    BST_SARC_hrs = [_[4] for _ in output.values()]

    df = pd.DataFrame({"Shifts": shifts, "Hours": hours, "Rates": rates, "Billable": billable,
                    "BST_ins": BST_insurance_hrs, "BST_SARC": BST_SARC_hrs})
    df["Shifts"] = df["Shifts"].astype(cat_size_order)
    df = df.sort_values("Shifts")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice to send to GT"

    # adjust column and row widths
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20

    heights = [16, 36, 36, 21, 21, 19, 21, 21, 15, 36, 17]
    heights += [21] * 24
    heights += [24, 16, 20, 15, 20]

    for _ in range(len(heights)):
        ws.row_dimensions[_+1].height = heights[_]

    # header
    ws.cell(row=2, column=5).value = "INVOICE"
    ws.cell(row=2, column=5).alignment = Alignment(horizontal="center",
                                                vertical="center")

    ws.cell(row=3, column=1).value = "Vendor:"
    ws.cell(row=4, column=1).value = "Client Name:"
    ws.cell(row=5, column=1).value = "Service Code:"
    ws.cell(row=3, column=3).value = "Nova Home Support"
    ws.cell(row=4, column=3).value = "M. Wohlmorantz"
    ws.cell(row=5, column=3).value = 320
    ws.cell(row=4, column=7).value = "Billed To:"
    ws.cell(row=5, column=7).value = "Pay To:"
    ws.cell(row=4, column=8).value = "GT Processing Dept"
    ws.cell(row=5, column=8).value = "Nova Home Support"

    today = datetime.date.today()
    today_date = today.strftime("%B %d %Y")
    today_date = today_date.split()
    curr_month = today_date[0]
    curr_date = int(today_date[1])
    curr_year = int(today_date[2])

    today = datetime.date.today()

    ws.cell(row=6, column=3).value = "(Drop down Menu)"
    ws.cell(row=7, column=1).value = "Date of"
    ws.cell(row=8, column=1).value = "Invoice:"
    ws.cell(row=7, column=2).value = "Month"
    ws.cell(row=8, column=2).value = curr_month
    ws.cell(row=7, column=3).value = "Day"
    ws.cell(row=8, column=3).value = curr_date
    ws.cell(row=7, column=4).value = "Year"
    ws.cell(row=8, column=4).value = curr_year
    ws.cell(row=7, column=5).value = "Period of invoice: " + PAY_PERIOD

    months = "January,February,March,April,May,June,July,August,September,October,November,December"
    dv = DataValidation(type="list", formula1="\""+months+"\"", allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["B8"])
    dv.add(ws["A12"])

    days = [str(_) for _ in range(1, 32)]
    days = ",".join(days)
    dv = DataValidation(type="list", formula1="\""+days+"\"", allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["C8"])

    years = [str(_) for _ in range(2021, 2027)]
    years = ",".join(years)
    dv = DataValidation(type="list", formula1="\""+years+"\"", allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["D8"])
    dv.add(ws["A13"])

    # table headers
    ws.cell(row=10, column=1).value = "Billing"
    ws.cell(row=10, column=5).value = "Original Gross Hours"
    ws.cell(row=10, column=5).alignment = Alignment(wrap_text=True)
    # col 6 (F): BST/BCBA Hours Billed to Insurance
    #   # RBT hours paid as BST1, # RBT hours paid as BST2, etc.
    ws.cell(row=10, column=6).value = "BST/BCBA Hours Billed to Insurance"
    ws.cell(row=10, column=6).alignment = Alignment(wrap_text=True)
    # col 7 (G): Hours Billed to SARC
    #   deduct col 6 from original gross hours
    ws.cell(row=10, column=7).value = "Hours Billed to SARC"
    ws.cell(row=10, column=7).alignment = Alignment(wrap_text=True)
    # col 8 (H): same
    ws.cell(row=10, column=8).value = "Rate"
    # col 9 (I): for BST, it is (col 7) * (col 8)
    ws.cell(row=10, column=9).value = "Amounts Billed to SARC"
    ws.cell(row=11, column=1).value = "Month of:"
    ws.cell(row=11, column=2).value = "Sub Code"
    ws.cell(row=11, column=3).value = "Nova code"
    ws.cell(row=12, column=1).value = curr_month
    ws.cell(row=13, column=1).value = curr_year

    # add data
    df = df.reset_index(drop=True)
    df.loc[len(df.index)] = ["","","","","",""]
    df.loc[len(df.index)] = ["Nova Leadership Costs", "", "", total_mgr, "", ""]
    temp = df["Billable"].replace("", 0)
    total = temp.sum()

    fmt_acct = u'_($* #,##0.00_);[Red]_($* (#,##0.00);_($* -_0_0_);_(@'
    ws.cell(row=11, column=7).fill = PatternFill(fgColor=Color("D9D9D9"), fill_type="solid")
    ws.cell(row=11, column=9).fill = PatternFill(fgColor=Color("A6A6A6"), fill_type="solid")

    for ind in df.index:
        if ind != len(df.index) - 2:
            ws.cell(row=12+ind, column=2).value = 320
        ws.cell(row=12+ind, column=3).value = df["Shifts"][ind]
        ws.cell(row=12+ind, column=5).value = df["Hours"][ind]
        # col 6 (F): BST/BCBA Hours Billed to Insurance
        #   # RBT hours paid as BST1, # RBT hours paid as BST2, etc.
        ws.cell(row=12+ind, column=6).value = df["BST_ins"][ind]
        # col 7 (G): Hours Billed to SARC
        #   deduct col 6 from original gross hours
        ws.cell(row=12+ind, column=7).value = df["BST_SARC"][ind]
        ws.cell(row=12+ind, column=7).fill = PatternFill(fgColor=Color("D9D9D9"), fill_type="solid")
        # col 8 (H): same
        ws.cell(row=12+ind, column=8).value = df["Rates"][ind]
        ws.cell(row=12+ind, column=8).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        # col 9 (I): for BST, it is (col 7) * (col 8)
        ws.cell(row=12+ind, column=9).value = df["Billable"][ind]
        ws.cell(row=12+ind, column=9).number_format = fmt_acct
        ws.cell(row=12+ind, column=9).fill = PatternFill(fgColor=Color("A6A6A6"), fill_type="solid")

    ws.cell(row=36, column=9).value = round(total, 2)
    ws.cell(row=36, column=9).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # footer
    ws.cell(row=36, column=2).value = "TOTAL"
    ws.cell(row=36, column=2).font = Font(bold=True)
    ws.cell(row=36, column=6).value = "Billing to GT Processing"
    ws.cell(row=38, column=2).value = "Signature"
    ws.cell(row=40, column=2).value = "Date:"
    ws.merge_cells("C38:G38")
    ws.merge_cells("C40:G40")
    ws.cell(row=38, column=3).value = "Alison Morantz"
    ws.cell(row=38, column=3).alignment = Alignment(horizontal="center")

    today = datetime.date.today()
    todayDate = today.strftime("%m.%d.%y")
    ws.cell(row=40, column=3).value = todayDate
    ws.cell(row=40, column=3).alignment = Alignment(horizontal="center")

    # fill colors
    # blue
    ws["A2"].fill = PatternFill(fgColor=Color("9BC2E6"), fill_type="solid")
    for col in "BCDEFGH":
        ws[col+"2"].fill = PatternFill(fgColor=Color("9BC2E6"), fill_type="solid")
        ws[col+"36"].fill = PatternFill(fgColor=Color("9BC2E6"), fill_type="solid")
    ws["I36"].fill = PatternFill(fgColor=Color("9BC2E6"), fill_type="solid")
    ws["A37"].fill = PatternFill(fgColor=Color("9BC2E6"), fill_type="solid")

    # light gray
    for row in range(3, 6):
        for col in "ABCDEFGHI":
            ws[col+str(row)].fill = PatternFill(fgColor=Color("F2F2F2"), fill_type="solid")

    # medium gray
    ws["I2"].fill = PatternFill(fgColor=Color("F2F2F2"), fill_type="solid")
    for row in range(6, 10):
        for col in "ABCDEFGHI":
            ws[col+str(row)].fill = PatternFill(fgColor=Color("D9D9D9"), fill_type="solid")
    ws["G10"].fill = PatternFill(fgColor=Color("D9D9D9"), fill_type="solid")

    # define border types
    allBorder = Border(top = Side(border_style='thin', color='FF000000'),
                    right = Side(border_style='thin', color='FF000000'),
                    bottom = Side(border_style='thin', color='FF000000'),
                    left = Side(border_style='thin', color='FF000000'))
    topLeft = Border(top = Side(border_style='thin', color='FF000000'),
                    left = Side(border_style='thin', color='FF000000'))
    topRight = Border(top = Side(border_style='thin', color='FF000000'),
                    right = Side(border_style='thin', color='FF000000'))
    bottomLeft = Border(bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
    bottomRight = Border(bottom = Side(border_style='thin', color='FF000000'),
                        right = Side(border_style='thin', color='FF000000'))
    top = Border(top = Side(border_style='thin', color='FF000000'))
    right = Border(right = Side(border_style='thin', color='FF000000'))
    bottom = Border(bottom = Side(border_style='thin', color='FF000000'))
    left = Border(left = Side(border_style='thin', color='FF000000'))

    # border: drop down menu
    for row in range(7, 9):
        for col in "BCD":
            ws[col+str(row)].border = allBorder
    ws["A7"].border = topLeft
    ws["A8"].border = bottomLeft
    ws["E7"].border = top
    ws["E8"].border = bottom
    ws["F7"].border = top
    ws["F8"].border = bottom
    ws["G7"].border = topRight
    ws["G8"].border = bottomRight
    ws["A6"].border = topLeft
    ws["I6"].border = topRight
    ws["I9"].border = bottomRight
    ws["A9"].border = bottomLeft
    for col in "BCDEFGH":
        ws[col+"6"].border = top
        ws[col+"9"].border = bottom
    ws["I7"].border = right
    ws["I8"].border = right

    # border: header
    ws["A2"].border = topLeft
    ws["A3"].border = topLeft
    ws["A4"].border = left
    ws["A5"].border = left
    ws["I3"].border = right
    ws["I4"].border = right
    ws["I5"].border = right
    for col in "BCDEFG":
        ws[col+"2"].border = top
        ws[col+"3"].border = top
    ws["H2"].border = Border(top = Side(border_style='medium', color='FF000000'),
                            right = Side(border_style='thin', color='FF000000'),
                            bottom = Side(border_style='medium', color='FF000000'),
                            left = Side(border_style='medium', color='FF000000'))
    ws["I2"].border = Border(top = Side(border_style='medium', color='FF000000'),
                            right = Side(border_style='medium', color='FF000000'),
                            bottom = Side(border_style='medium', color='FF000000'))

    # border: main content
    ws["B11"].border = topLeft
    ws["H11"].border = topRight
    ws["B35"].border = bottomLeft
    ws["H35"].border = bottomRight

    # horizontal lines
    for col in "CDEFG":
        ws[col+"11"].border = top
        ws[col+"35"].border = bottom
        ws[col+"38"].border = bottom
        ws[col+"40"].border = bottom
    for col in "BCDEFGH":
        ws[col+"36"].border = bottom

    # vertical lines
    ws["A10"].border = left
    ws["A11"].border = left
    ws["A35"].border = left
    ws["A36"].border = left
    ws["A37"].border = bottomLeft
    ws["I10"].border = right
    ws["I11"].border = topRight
    ws["I35"].border = bottomRight
    ws["I36"].border = bottomRight
    for row in range(12, 35):
        ws["A"+str(row)].border = left
        ws["B"+str(row)].border = left
        ws["H"+str(row)].border = right
        ws["I"+str(row)].border = right

    # font
    ws['E2'].font = Font(name="Aparajita", size=24)
    for row in range(3, 6):
        for col in "AG":
            ws[col+str(row)].font = Font(name="Aparajita", size=14, bold=True)
        for col in "BCDEF":
            ws[col+str(row)].font = Font(name="Aparajita", size=14)
    ws['A7'].font = Font(name="Aparajita", size=14, bold=True)
    ws['A8'].font = Font(name="Aparajita", size=14, bold=True)
    ws['B7'].font = Font(name="Aparajita", size=14, bold=True)
    ws['B8'].font = Font(name="Aparajita", size=14)
    ws['C7'].font = Font(name="Aparajita", size=14, bold=True)
    ws['C8'].font = Font(name="Aparajita", size=14)
    ws['D7'].font = Font(name="Aparajita", size=14, bold=True)
    ws['D8'].font = Font(name="Aparajita", size=14)

    for row in range(12, 36):
        for col in "BCDEFGH":
            ws[col+str(row)].font = Font(name="Aparajita", size=14)

    # row 10: bold from A to I
    for col in "ABCDEFGHI":
        ws[col+"10"].font = Font(name="Aparajita", bold=True)
    # row 11: bold in A; apply font from B to I
    ws["A11"].font = Font(name="Aparajita", bold=True)
    for col in "BCDEFGHI":
        ws[col+"11"].font = Font(name="Aparajita")
    # A12: bold
    ws["A12"].font = Font(name="Aparajita", bold=True)
    # col A, rows 13-36: apply font
    for row in range(13, 37):
        ws["A"+str(row)].font = Font(name="Aparajita")
    # row 36, col B to H: apply font
    for col in "CDEFGH":
        ws[col+"36"].font = Font(name="Aparajita")
    # I36: font size 16
    ws["I36"].font = Font(name="Aparajita", size=16)
    ws["C6"].font = Font(size=8)
    ws["E7"].font = Font(size=10)
    ws["H4"].font = Font(size=14)
    ws["H5"].font = Font(size=12)
    ws["C38"].font = Font(name="Brush Script MT", size=14)
    ws["C40"].font = Font(name="Brush Script MT", size=14)
    wb.save(save_path)
    wb.close()

    book = load_workbook(save_path)
    writer = pd.ExcelWriter(save_path, engine='openpyxl') 
    writer.book = book

    df_benefits.to_excel(writer, "Nova Leadership Cost Breakdowns", index=False)

    df_shift_merged['Holiday Worked Duration (Hours)'] = (df_shift_merged['Holiday Worked Duration (Minutes)']/60).round(2)
    df_shift_merged['Hrs. Worked'] = df_shift_merged['Min. Worked']/60
    df_shift_merged = df_shift_merged.round(decimals=2)
    df_shift_merged['Day of the Week'] = df_shift_merged['CIDT'].dt.day_name()
    df_shift_merged = df_shift_merged[['Name', 'First Name', 'Last Name', 'Shift_original', 'Shift','Day of the Week', 'Check-In Date', 
                                       'Check-In Time', 'Check-Out Date', 'Check-Out Time', 'Min. Worked', 'Hrs. Worked',  'Regular Hourly Wage', 
                                       'BOT Hourly Wage', 'Accrual Rate','Error Check', 'CIDT', 'CODT','Holiday Worked Duration (Minutes)',
                                       'Holiday Worked Duration (Hours)']]
    df_shift_merged = df_shift_merged.reindex(columns=['Name', 'First Name', 'Last Name', 'Shift_original', 'Shift','Day of the Week', 'Check-In Date', 
                                                       'Check-In Time', 'Check-Out Date', 'Check-Out Time', 'Min. Worked', 'Hrs. Worked',  
                                                       'Regular Hourly Wage', 'BOT Hourly Wage', 'Accrual Rate','Error Check', 'CIDT', 
                                                       'CODT','Holiday Worked Duration (Minutes)','Holiday Worked Duration (Hours)'])
    df_shift_merged.to_excel(writer, "Shift Breakdowns", index=False)

    writer.save()

    wb = load_workbook(save_path)
    ws1 = wb["Nova Leadership Cost Breakdowns"]
    ws1.cell(row=len(df_benefits.index)+4, column=1).value = "MANAGERS' TOTAL WAGES & BENEFITS"
    ws1.cell(row=len(df_benefits.index)+5, column=1).value = total_mgr
    ws1.cell(row=len(df_benefits.index)+4, column=1).font = Font(bold=True)
    for column_cells in ws1.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws1.column_dimensions[new_column_letter].width = new_column_length*1.23

    ws2 = wb["Shift Breakdowns"]
    for column_cells in ws2.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws2.column_dimensions[new_column_letter].width = new_column_length*1.23

    wb.save(save_path)
    wb.close()
    # return the underlying dataset.
    return df

def output_underlying(mgr_pr, non_mgr_pr, invoice_df, save_path, PAY_PERIOD, FULL_CYCLE):
    '''
    Output the underlying payroll information as a square dataset in csv format
    
    mgr_pr -- manager's payroll dictionary 
    non_mgr_pr -- non-managers's payroll dictionary
    invoice_df -- pandas dataframe containing info about invoice
    FULL_CYCLE -- boolean: if we are processing a full pay cycle
    '''
    #the underlying payroll information
    noumenon = pd.DataFrame(columns=['Pay Period','Name', 'Total Gross Wage', 'Shift', 'Min. Worked', 'Hrs. Worked', 'Wage', 'Gross Wages', 
                                    'Hrs. YTD', 'Hrs. Worked This Period','Hire Date', 'Calendar Days Since Hire Date', 
                                    'Vac. Accrued YTD','Vac. Taken YTD', 'Vac. Accrued This Period', 'Vac. Taken This Period',
                                    'Vac. Balance', 'Sick Bank YTD', 'Sick Taken YTD','Sick Taken This Period', 'Sick Balance'])
    #flatten in a loop
    for pack in [*mgr_pr, *non_mgr_pr]:
        payroll = pack['payroll']
        if len(payroll) == 0:
            continue
        summary =  pd.concat([pack['summary']] * len(payroll), ignore_index=True)
        acc_A = pd.concat([pack['accrued_A']] * len(payroll), ignore_index=True)
        acc_B = pd.concat([pack['accrued_B']] * len(payroll), ignore_index=True)
        acc_C = pd.concat([pack['accrued_C']] * len(payroll), ignore_index=True)
        flattened = pd.concat([payroll, summary, acc_A, acc_B, acc_C], axis=1)
        noumenon = pd.concat([noumenon, flattened], axis=0, ignore_index=True)
    #save file
    if FULL_CYCLE:
        #drop manager only column
        noumenon = noumenon.drop(['Total Hours Worked'], axis=1)
        #add pay period to invoice
        invoice_df['Pay Period'] = PAY_PERIOD
        #reorganized order of cols
        invoice_df = invoice_df[['Pay Period'] + [col for col in invoice_df if col != 'Pay Period']]
        invoice_df = invoice_df.rename(columns={"Shifts": "Shift",
                                                "Hours": "Gross_Hrs",
                                                "Rates": "Rate",
                                                "Billable": "SARC_Billed_Amt",
                                                "BST_ins": "Ins_Billed_Hrs",
                                                "BST_SARC": "SARC_Billed_Hrs"})
        with pd.ExcelWriter(save_path + "/" + f"MACHINE_READABLE_OUTPUT - {PAY_PERIOD}.xlsx") as writer:
            # Write each dataframe to a separate tab
            noumenon.to_excel(writer, sheet_name='Payroll', index=False)
            invoice_df.to_excel(writer, sheet_name='Invoice', index=False)
    else:
        with pd.ExcelWriter(save_path + "/" + f"MACHINE_READABLE_OUTPUT - {PAY_PERIOD}.xlsx") as writer:
            noumenon.to_excel(writer, sheet_name='Payroll', index=False)


# Copyright (c) [2023] [Nova Home Support LLC]
# This code is licensed under the Creative Commons Attribution-NonCommercial 4.0 International License. See LICENSE.md for details.
